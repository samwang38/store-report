#!/usr/bin/env python3
"""門市報表製作 — 本機 Web 工具

可下拉選任一門市，產生該門市的士林式 13 張週報（含個人項目動態填入）。
EPB 即時查詢免登入，採用「選日期 → 背景產生 → 下載」的非同步流程。

報表核心（EPB 查詢 / 13 張填表 / 個人項目動態填入）移植自 live-report-app，
邏輯零改動；本檔僅去掉登入閘、改成非同步任務佇列、加上免登入門市下拉。
"""
import copy
import csv
import importlib.util
import json
import os
import subprocess
import sys
import threading
import time
import traceback
import urllib.parse
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd

try:
    import shoppertrak_traffic as traffic_mod
except Exception as _exc:  # 缺套件等情況下仍讓 server 正常啟動
    traffic_mod = None
    _TRAFFIC_IMPORT_ERROR = str(_exc)

try:
    import dss_client as dss_mod
except Exception as _exc:
    dss_mod = None
    _DSS_IMPORT_ERROR = str(_exc)


ROOT = Path(__file__).resolve().parent
STATIC_ROOT = ROOT / "static"
TEMPLATE_PATH = ROOT / "週報模板.xlsx"
DATA_DIR = ROOT / "銷售資料"
ENGINE_PATH = ROOT / "fill_weekly_excel.py"
SACARE_PATH = DATA_DIR / "SAcare對應價目表.xlsx"

# Java 路徑：環境變數 EPB_JAVA_HOME 優先（指到 JDK Home），否則沿用預設 JDK 1.8 安裝路徑
_DEFAULT_JDK = "/Library/Java/JavaVirtualMachines/jdk1.8.0_251.jdk/Contents/Home"
_JAVA_HOME = os.environ.get("EPB_JAVA_HOME", _DEFAULT_JDK)
JAVA = (
    f"{_JAVA_HOME}/jre/bin/java"
    if Path(f"{_JAVA_HOME}/jre/bin/java").exists()
    else f"{_JAVA_HOME}/bin/java"
)
JAVAC = f"{_JAVA_HOME}/bin/javac"
JAVA_CP = f"{ROOT}:/Library/EPBrowser/EPB/Shell/lib/*:/Library/EPBrowser/EPB/Shell/shell.jar"

LOCAL_CONFIG_PATH = ROOT / "local_config.json"


def _local_config():
    try:
        return json.loads(LOCAL_CONFIG_PATH.read_text("utf-8"))
    except Exception:
        return {}


# EPB WSDL 端點：環境變數 EPB_WSDL_URL → local_config.json 的 epb.wsdlUrl → 預設內網位址
EPB_WSDL_URL = (
    os.environ.get("EPB_WSDL_URL")
    or _local_config().get("epb", {}).get("wsdlUrl")
    or "http://192.168.1.177:8080/EPB_AP_EPB/EPB_AP?wsdl"
)

ORG_ID = "01"
DEFAULT_SHOP_ID = "004"

# EPB 門市主檔查詢失敗時的後備清單（取自北一區 multistore_engine 的 STORES）
FALLBACK_STORES = {
    "004": "士林門市",
    "005": "微風門市",
    "024": "美麗華門市",
    "046": "阿波羅門市",
    "054": "大葉高島屋門市",
    "057": "羅東門市",
    "068": "新店裕隆城",
}

REPORT_SHEETS = [
    "1.主機銷售台數",
    "2.門市週報 ",
    "3.3PP配件比較",
    "4.3PP 銷售排名",
    "5.VAP銷售排名",
    "6.搭售統計_週",
    "7.搭售統計_月",
    "8.個人新制獎金",
    "9.個人週主機",
    "10.個人月主機",
    "11.個人月3PP",
    "12.月報YOY",
    "13.3PP YOY",
]

TRANS_TYPE_MAP = {
    "A": "銷售",
    "E": "銷退",
    "G": "訂金",
    "H": "尾款",
    "J": "退訂",
}

STANDARD_COLUMNS = [
    "交易類型",
    "單據日期",
    "單據代碼",
    "員工代碼",
    "員工名稱",
    "存貨代碼",
    "名稱",
    "數量",
    "銷售金額(含稅)",
    "銷退金額",
    "單位成本",
    "品牌代碼",
    "類別1代碼",
    "類別3代碼",
    "類別4代碼",
    "類別6代碼",
    "折扣",
    "等級代碼",
    "淨銷售金額(未稅)",
]


def load_engine():
    if not ENGINE_PATH.exists():
        raise RuntimeError(f"找不到週報引擎：{ENGINE_PATH}")
    spec = importlib.util.spec_from_file_location("shilin_weekly_engine", ENGINE_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


engine = load_engine()

JOBS = {}
_LOCK = threading.Lock()


# ─── 共用小工具（移植自 live-report-app）────────────────────────────────
def json_default(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def safe_shop_id(raw):
    value = str(raw or DEFAULT_SHOP_ID).strip().upper()
    if not value:
        return DEFAULT_SHOP_ID
    if not all(ch.isalnum() or ch in "_-" for ch in value):
        raise ValueError("門市代碼格式錯誤")
    return value


def quote_sql(value):
    return "'" + str(value).replace("'", "''") + "'"


def parse_date(value, fallback=None):
    if not value:
        if fallback is None:
            raise ValueError("缺少日期")
        return fallback
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def last_saturday(as_of=None):
    as_of = as_of or date.today()
    days = (as_of.weekday() + 2) % 7  # Sat→0, Sun→1, Mon→2, …
    return as_of - timedelta(days=days if days else 7)


def compute_periods(wk_start, wk_end):
    year, prev_year = wk_start.year, wk_start.year - 1
    cross_month = wk_end.month != wk_start.month
    report_month = wk_start.month
    prev_month = report_month - 1 if report_month > 1 else 12
    prev_mo_year = year if report_month > 1 else year - 1

    if cross_month:
        mo_start = date(year, report_month, 1)
        mo_end = date(year, report_month, engine.monthrange(year, report_month)[1])
        lm_start = date(prev_mo_year, prev_month, 1)
        lm_end = date(prev_mo_year, prev_month, engine.monthrange(prev_mo_year, prev_month)[1])
        ly_start = date(prev_year, report_month, 1)
        ly_end = date(prev_year, report_month, engine.monthrange(prev_year, report_month)[1])
    else:
        mo_start = date(year, report_month, 1)
        mo_end = wk_end
        lm_last = engine.monthrange(prev_mo_year, prev_month)[1]
        lm_start = date(prev_mo_year, prev_month, 1)
        lm_end = date(prev_mo_year, prev_month, min(wk_end.day, lm_last))
        ly_last = engine.monthrange(prev_year, report_month)[1]
        ly_start = date(prev_year, report_month, 1)
        ly_end = date(prev_year, report_month, min(wk_end.day, ly_last))

    prev_wk_end = wk_start - timedelta(days=1)
    prev_wk_start = prev_wk_end - timedelta(days=6)
    return {
        "wk_start": wk_start,
        "wk_end": wk_end,
        "prev_wk_start": prev_wk_start,
        "prev_wk_end": prev_wk_end,
        "mo_start": mo_start,
        "mo_end": mo_end,
        "lm_start": lm_start,
        "lm_end": lm_end,
        "ly_start": ly_start,
        "ly_end": ly_end,
        "ytd_cur_start": date(year, 1, 1),
        "ytd_prv_start": date(prev_year, 1, 1),
    }


# 進階「自訂計算區間」：前端欄位名 → dates 鍵（本週 wkStart/wkEnd 在 compute_periods 前已套用為基準）
PERIOD_OVERRIDE_KEYS = {
    "prevWkStart": "prev_wk_start",
    "prevWkEnd": "prev_wk_end",
    "moStart": "mo_start",
    "moEnd": "mo_end",
    "lmStart": "lm_start",
    "lmEnd": "lm_end",
    "lyStart": "ly_start",
    "lyEnd": "ly_end",
}

# (起鍵, 迄鍵, 顯示名) 供區間合法性檢查
PERIOD_PAIRS = [
    ("prev_wk_start", "prev_wk_end", "上週"),
    ("wk_start", "wk_end", "本週"),
    ("mo_start", "mo_end", "本月"),
    ("lm_start", "lm_end", "上月"),
    ("ly_start", "ly_end", "去年同期"),
]


def apply_period_overrides(dates, overrides, log=lambda m: None):
    """套用進階面板的自訂區間（提供才覆寫），並驗證各組起訖合法。"""
    applied = []
    for okey, dkey in PERIOD_OVERRIDE_KEYS.items():
        raw = overrides.get(okey)
        if raw:
            dates[dkey] = parse_date(raw)
            applied.append(dkey)
    for skey, ekey, label in PERIOD_PAIRS:
        if dates[ekey] < dates[skey]:
            raise ValueError(f"{label}區間的結束日早於起始日")
    if applied:
        log(f"套用自訂區間：上週 {dates['prev_wk_start']}~{dates['prev_wk_end']}、"
            f"本月 {dates['mo_start']}~{dates['mo_end']}、上月 {dates['lm_start']}~{dates['lm_end']}、"
            f"去年同期 {dates['ly_start']}~{dates['ly_end']}")


def compile_java(source_name):
    source = ROOT / source_name
    target = ROOT / source_name.replace(".java", ".class")
    if target.exists() and target.stat().st_mtime >= source.stat().st_mtime:
        return
    proc = subprocess.run(
        [JAVAC, "-cp", JAVA_CP, str(source)],
        cwd=str(ROOT),
        text=True,
        capture_output=True,
        timeout=30,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or proc.stdout.strip())


def run_remote(sql, timeout=180, max_rows=100000):
    compile_java("EPBReportQuery.java")
    proc = subprocess.run(
        [
            JAVA,
            "-Dsun.net.client.defaultConnectTimeout=5000",
            "-Dsun.net.client.defaultReadTimeout=120000",
            f"-Depb.wsdl={EPB_WSDL_URL}",
            "-cp",
            JAVA_CP,
            "EPBReportQuery",
            sql,
            str(max_rows),
        ],
        cwd=str(ROOT),
        text=True,
        capture_output=True,
        timeout=timeout,
    )
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout).strip())
    lines = [line for line in proc.stdout.splitlines() if line.strip()]
    if not lines:
        return [], []
    reader = csv.reader(lines, delimiter="\t")
    rows = list(reader)
    headers = rows[0]
    return headers, rows[1:]


def number(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def text(value):
    return "" if value is None else str(value).strip()


def empty_standard_df():
    df = pd.DataFrame(columns=STANDARD_COLUMNS)
    df["NET"] = pd.Series(dtype=float)
    return df


def standardize_remote_records(records):
    rows = []
    for rec in records:
        trans_code = text(rec.get("TRANS_TYPE")).upper()
        doc_date = pd.to_datetime(text(rec.get("DOC_DATE"))[:19], errors="coerce")
        qty = number(rec.get("STK_QTY"))
        line_total_net = number(rec.get("LINE_TOTAL_NET"))
        line_tax = number(rec.get("LINE_TAX"))
        gross_amount = line_total_net + line_tax
        unit_cost = number(rec.get("TRN_COST_PRICE")) or number(rec.get("COST_PRICE"))
        rows.append(
            {
                "交易類型": TRANS_TYPE_MAP.get(trans_code, trans_code or "銷售"),
                "單據日期": doc_date,
                "單據代碼": text(rec.get("DOC_ID")),
                "員工代碼": text(rec.get("EMP_ID1")),
                "員工名稱": text(rec.get("EMP_NAME")),
                "存貨代碼": text(rec.get("STK_ID")),
                "名稱": text(rec.get("NAME")),
                "數量": qty,
                "銷售金額(含稅)": gross_amount,
                "銷退金額": 0.0,
                "單位成本": unit_cost,
                "品牌代碼": number(rec.get("BRAND_ID")),
                "類別1代碼": number(rec.get("CAT1_ID")),
                "類別3代碼": number(rec.get("CAT3_ID")),
                "類別4代碼": number(rec.get("CAT4_ID")),
                "類別6代碼": number(rec.get("CAT6_ID")),
                "折扣": number(rec.get("DISC_NUM")),
                "等級代碼": "",
                "淨銷售金額(未稅)": line_total_net,
            }
        )
    if not rows:
        return empty_standard_df()
    df = pd.DataFrame(rows, columns=STANDARD_COLUMNS)
    df["單據日期"] = pd.to_datetime(df["單據日期"])
    df["存貨代碼"] = df["存貨代碼"].astype(str).str.strip()
    df["員工代碼"] = df["員工代碼"].astype(str).str.strip()
    df["等級代碼"] = df["等級代碼"].astype(str).str.strip()
    for col in ["品牌代碼", "類別1代碼", "類別3代碼", "類別4代碼", "類別6代碼"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["NET"] = df["銷售金額(含稅)"].fillna(0) + df["銷退金額"].fillna(0)
    return df


def remote_sales_df(shop_id, ranges):
    """ranges: [(start_date, end_date), ...]。多組區間以 OR 合併成單一 SQL，
    只啟動一次 JVM / 一次 SOAP 來回，由呼叫端再依日期切分。"""
    ranges = [(s, e) for s, e in ranges if e >= s]
    if not ranges:
        return empty_standard_df()
    date_conds = " or ".join(
        f"(l.doc_date >= to_date({quote_sql(s.isoformat())}, 'yyyy-mm-dd')"
        f" and l.doc_date < to_date({quote_sql((e + timedelta(days=1)).isoformat())}, 'yyyy-mm-dd'))"
        for s, e in ranges
    )
    sql = f"""
select
  l.trans_type,
  l.doc_date,
  l.doc_id,
  l.emp_id1,
  coalesce(e.name, l.emp_id1) as emp_name,
  l.emp_id2,
  l.stk_id,
  l.name,
  l.stk_qty,
  l.line_total_net,
  l.line_tax,
  l.line_total,
  l.cost_price,
  l.trn_cost_price,
  l.brand_id,
  l.cat1_id,
  l.cat3_id,
  l.cat4_id,
  l.cat6_id,
  l.disc_num,
  l.line_no
from poslinev_bi l
left join ep_user e on e.user_id = l.emp_id1
where l.org_id = {quote_sql(ORG_ID)}
  and l.shop_id = {quote_sql(shop_id)}
  and ({date_conds})
order by l.doc_date, l.doc_id, l.line_no
"""
    headers, rows = run_remote(sql, max_rows=500000)
    records = []
    for row in rows:
        row = (row + [""] * len(headers))[: len(headers)]
        records.append({headers[i].upper(): row[i] for i in range(len(headers))})
    return standardize_remote_records(records)


_SACARE_CACHE = {"mtime": None, "prices": None}


def load_sacare_cached():
    """SAcare 價目表以檔案 mtime 快取，檔案沒變就不重讀 Excel。"""
    mtime = SACARE_PATH.stat().st_mtime
    if _SACARE_CACHE["mtime"] != mtime:
        _SACARE_CACHE["prices"] = engine.load_sacare(SACARE_PATH)
        _SACARE_CACHE["mtime"] = mtime
    return _SACARE_CACHE["prices"]


def load_epb_data(shop_id, dates, quarter_start):
    sacare_prices = load_sacare_cached()
    # 納入 Sheet 12/13 年對年區間（截止日可由前端自訂，可能晚於週結束日）
    yoy_cur_s, yoy_cur_e, yoy_prv_s, yoy_prv_e = engine._yoy_periods(dates)
    # df_cur 涵蓋上週/本週/本月/上月，df_prev 涵蓋去年同期；納入自訂覆寫區間邊界以確保資料齊全
    cur_start = min(dates["ytd_cur_start"], dates["prev_wk_start"], dates["wk_start"],
                    dates["mo_start"], dates["lm_start"], quarter_start, yoy_cur_s)
    cur_end = max(dates["mo_end"], dates["wk_end"], dates["prev_wk_end"], dates["lm_end"], yoy_cur_e)
    prv_start = min(dates["ytd_prv_start"], dates["ly_start"], yoy_prv_s)
    prv_end = max(dates["ly_end"], yoy_prv_e)
    # 兩個年度區間合併成單一查詢（單次 JVM/SOAP），回來後依日期切分
    df_all = remote_sales_df(shop_id, [(cur_start, cur_end), (prv_start, prv_end)])
    dts = pd.to_datetime(df_all["單據日期"], errors="coerce")
    df_cur = df_all[(dts >= pd.Timestamp(cur_start)) & (dts < pd.Timestamp(cur_end + timedelta(days=1)))].reset_index(drop=True)
    df_prev = df_all[(dts >= pd.Timestamp(prv_start)) & (dts < pd.Timestamp(prv_end + timedelta(days=1)))].reset_index(drop=True)
    return df_cur, df_prev, sacare_prices, {
        "shopId": shop_id,
        "currentRange": f"{cur_start.isoformat()}~{cur_end.isoformat()}",
        "previousRange": f"{prv_start.isoformat()}~{prv_end.isoformat()}",
        "currentRows": int(len(df_cur)),
        "previousRows": int(len(df_prev)),
    }


# ─── ShopperTrak 來客數（人流）─────────────────────────────────────────
def _emp_count(payload):
    v = payload.get("employeeCount", payload.get("employee_count"))
    try:
        return int(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def load_traffic(shop_id, dates, log=lambda m: None):
    """回傳 (traffic_sheet2, traffic_sheet10)。
    任何失敗（未裝套件 / 未設帳密 / 無 siteId / 登入或查詢錯誤）→ 回 ({}, {})，
    並記錄 log，讓報表「略過來客數、其餘照常產生」。"""
    if traffic_mod is None:
        log("未載入來客數模組，略過來客數")
        return {}, {}
    if not traffic_mod.has_credentials():
        log("未設定 ShopperTrak 帳密，略過來客數")
        return {}, {}
    if not traffic_mod.site_id_for_shop(shop_id):
        log(f"門市 {shop_id} 無對應 ShopperTrak siteId，略過來客數")
        return {}, {}

    periods2 = {
        "上週": (dates["prev_wk_start"], dates["prev_wk_end"]),
        "本週": (dates["wk_start"], dates["wk_end"]),
        "本月": (dates["mo_start"], dates["mo_end"]),
        "上月": (dates["lm_start"], dates["lm_end"]),
        "去年": (dates["ly_start"], dates["ly_end"]),
    }
    yoy_cur_s, yoy_cur_e, yoy_prv_s, yoy_prv_e = engine._yoy_periods(dates)
    periods10 = {"cur": (yoy_cur_s, yoy_cur_e), "prv": (yoy_prv_s, yoy_prv_e)}

    try:
        tasks = [("t2", name, s, e) for name, (s, e) in periods2.items()]
        tasks += [("t10", name, s, e) for name, (s, e) in periods10.items()]
        results = {}
        # 第一個區間先單獨查（觸發登入、建立 token 快取），其餘並行查詢
        first = tasks[0]
        results[first[:2]] = traffic_mod.get_traffic_total(shop_id, first[2], first[3], log=log)
        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = {
                pool.submit(traffic_mod.get_traffic_total, shop_id, s, e, log=log): (which, name)
                for which, name, s, e in tasks[1:]
            }
            for fut, key in futures.items():
                results[key] = fut.result()
        t2 = {name: results[("t2", name)] for name in periods2}
        t10 = {name: results[("t10", name)] for name in periods10}
        log(f"  來客數：本週 {t2.get('本週'):,} / 本月 {t2.get('本月'):,}")
        return t2, t10
    except Exception as exc:
        log(f"  來客數查詢失敗，已略過：{exc}")
        return {}, {}


# ─── DSS 搭售統計（Sheet 6/7）─────────────────────────────────────────
# Sheet 6「搭售統計_週」：本期＝本週（週日～週六），歷史 3 列＝前 3 個完整週（W{財務季週次}）
# Sheet 7「搭售統計_月」：本期＝本月 1 日～週結束日，歷史 3 列＝前 3 個完整日曆月（{n}月）
# 定義（2026-06 與 DSS「3PP搭售率報表(人)」逐格核對確認）：
#   m=搭售台數（有搭配件的主機）、s=配件數、ms=零搭售台數
#   搭售率=s/m、零搭售比例=ms/(m+ms)、AA 3PP搭售率=Σs/Σm

BUNDLE_SHEET_WEEK = "6.搭售統計_週"
BUNDLE_SHEET_MONTH = "7.搭售統計_月"
# 各機種欄位起點（B/G/L/Q/V），每組 5 欄：台數、配件、搭售率、零搭售、零搭售比例
BUNDLE_GROUP_COLS = (("cpu", 2), ("iphone", 7), ("ipad", 12),
                     ("watch", 17), ("airpods", 22))
BUNDLE_3PP_COL = 27  # AA


def load_bundle_stats(shop_id, wk_start, wk_end, fiscal_start, log=lambda m: None):
    """抓搭售統計（8 個區間）。預設 EPB 直接計算（免登入），失敗自動退 DSS。

    local_config.json 可設 "bundleSource": "dss" 強制改回 DSS 為主。
    EPB 演算法與 DSS 經 5 週 155 數據點驗證（154 吻合，唯一差異為 DSS 端漏單）。
    失敗回 None → 略過 Sheet 6/7。
    """
    periods = {("w", 0): (wk_start, wk_end)}
    for i in (1, 2, 3):
        s = wk_start - timedelta(days=7 * i)
        periods[("w", i)] = (s, s + timedelta(days=6))
    mo_first = wk_end.replace(day=1)
    periods[("m", 0)] = (mo_first, wk_end)
    cursor = mo_first
    for i in (1, 2, 3):
        last_end = cursor - timedelta(days=1)
        periods[("m", i)] = (last_end.replace(day=1), last_end)
        cursor = last_end.replace(day=1)

    # bundleSource=dss → 強制只用 DSS（不退回 EPB）；預設 EPB 為主、DSS 備援
    forced_dss = bool(traffic_mod and traffic_mod.load_config().get("bundleSource") == "dss")
    order = ["dss"] if forced_dss else ["epb", "dss"]

    def via_epb():
        span_lo = min(s for s, _ in periods.values())
        span_hi = max(e for _, e in periods.values())
        prefetched = _epb_bundle_query(shop_id, span_lo, span_hi)
        return {key: epb_bundle_stats(shop_id, s, e, _prefetched=prefetched)
                for key, (s, e) in periods.items()}

    def via_dss():
        if dss_mod is None:
            raise RuntimeError("DSS 模組未載入")
        if not dss_mod.is_logged_in():
            raise RuntimeError("DSS 未登入")
        results = {}
        items = list(periods.items())
        # 第一個區間先單獨查（確認 session / 取 CSRF），其餘並行
        first_key, (fs, fe) = items[0]
        results[first_key] = dss_mod.fetch_bundle_stats(shop_id, fs, fe, log=log)
        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = {pool.submit(dss_mod.fetch_bundle_stats, shop_id, s, e): key
                       for key, (s, e) in items[1:]}
            for fut, key in futures.items():
                results[key] = fut.result()
        return results

    results = None
    for source in order:
        try:
            results = via_epb() if source == "epb" else via_dss()
            log(f"  搭售統計來源：{source.upper()}{'（強制 DSS）' if forced_dss else ''}")
            break
        except Exception as exc:
            log(f"  搭售統計 {source.upper()} 查詢失敗：{exc}")
    if results is None:
        log("  ✗ 搭售統計無法取得，略過 Sheet 6/7"
            + ("（強制 DSS 模式：請先登入 DSS 或取消強制）" if forced_dss else ""))
        return None

    def label_week(idx):
        s = periods[("w", idx)][0]
        if fiscal_start:
            _, w_num = engine.compute_fiscal_week_number(s, fiscal_start)
            return f"W{w_num}"
        return f"前{idx}週"

    return {
        "week": {
            "current": results[("w", 0)],
            "history": [(label_week(i), results[("w", i)]) for i in (1, 2, 3)],
        },
        "month": {
            "current": results[("m", 0)],
            "history": [(f"{periods[('m', i)][0].month}月", results[("m", i)])
                        for i in (1, 2, 3)],
        },
    }


def _bundle_totals(rows):
    tot = {k: {"m": 0, "s": 0, "ms": 0} for k, _ in BUNDLE_GROUP_COLS}
    for r in rows:
        for k in tot:
            for f in ("m", "s", "ms"):
                tot[k][f] += (r.get(k) or {}).get(f, 0)
    return tot


def _write_bundle_row(ws, row, label, stats):
    ws.cell(row=row, column=1).value = label
    sum_m = sum_s = 0
    for k, base in BUNDLE_GROUP_COLS:
        g = stats.get(k) or {"m": 0, "s": 0, "ms": 0}
        m, s, ms = g["m"], g["s"], g["ms"]
        sum_m += m
        sum_s += s
        ws.cell(row=row, column=base).value = m
        ws.cell(row=row, column=base + 1).value = s
        ws.cell(row=row, column=base + 2).value = round(s / m, 2) if m else None
        ws.cell(row=row, column=base + 3).value = ms
        ws.cell(row=row, column=base + 4).value = round(ms / (m + ms), 2) if (m + ms) else None
    ws.cell(row=row, column=BUNDLE_3PP_COL).value = round(sum_s / sum_m, 2) if sum_m else None


def _fill_bundle_sheet(ws, data, employees):
    """employees=[(code, name)]；員工列依報表員工順序，配 DSS emp_id 對應資料。"""
    by_id = {r["empId"]: r for r in data["current"]}
    matched = set()
    rows = []
    for code, name in employees:
        r = by_id.get(code)
        if r:
            matched.add(code)
        dss_name = (r or {}).get("empName") or ""
        # EPB 來源的 empName=員工代碼，改用報表員工名
        display = dss_name if (dss_name and dss_name != code) else name
        rows.append((display, r or {}))
    # DSS 有、但不在報表員工清單者也補列出，避免漏資料
    for r in data["current"]:
        if r["empId"] not in matched and any((r.get(k) or {}).get(f) for k, _ in BUNDLE_GROUP_COLS for f in ("m", "s", "ms")):
            rows.append((r.get("empName") or r["empId"], r))

    # 調整員工列數：模板第 2 列起至 Total 前一列
    total_row = next(
        (r for r in range(2, ws.max_row + 1) if str(ws.cell(row=r, column=1).value).strip() == "Total"),
        None,
    )
    if total_row is None:
        raise RuntimeError(f"{ws.title}：找不到 Total 列")
    slots = total_row - 2
    if len(rows) > slots:
        ws.insert_rows(total_row, len(rows) - slots)
        for r in range(total_row, total_row + len(rows) - slots):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c)._style = copy.copy(ws.cell(row=2, column=c)._style)
        total_row += len(rows) - slots
    elif len(rows) < slots:
        ws.delete_rows(2 + len(rows), slots - len(rows))
        total_row -= slots - len(rows)

    for i, (display, r) in enumerate(rows):
        _write_bundle_row(ws, 2 + i, display, r)
    _write_bundle_row(ws, total_row, "Total", _bundle_totals(data["current"]))
    for j, (label, hist_rows) in enumerate(data["history"]):
        _write_bundle_row(ws, total_row + 1 + j, label, _bundle_totals(hist_rows))


def fill_bundle_sheets(wb, bundle, employees, log=lambda m: None):
    _fill_bundle_sheet(wb[BUNDLE_SHEET_WEEK], bundle["week"], employees)
    _fill_bundle_sheet(wb[BUNDLE_SHEET_MONTH], bundle["month"], employees)


# ─── EPB 並行計算搭售統計（反推 DSS 規則，2026-06 當週 231 列明細逐列驗證一致）──
# 規則（與 DSS「3PP搭售率報表」完全重現）：
#   主機：C3=3001 且 C4∈{4001,4002}=CPU / 4004=iPhone / {4005,4006,4041}=iPad / 4038=Watch；
#         AirPods＝C3=3002 且 C4=4014 且 C6∈C6_AIRPODS
#   配件：C3=3003（3PP）。歸類：自身 C4 類別若在同單主機類別中→該類；
#         否則跟著同單唯一主機類；無主機單→跟著同 VIP 當日主機類（VIP 當日無主機→不計）
#   搭售：每單每類別有歸類配件時，該類「1 台」記搭售、其餘記零搭售（POS 組合販賣行為）；
#         純配件單可讓同 VIP 同日同類別尚無搭售的 1 台零搭售升級為搭售
#   交易：A/H 正計、E 負計（qty 相加自然沖銷）

BUNDLE_EPB_OWN_C4 = {4007.0: "cpu", 4009.0: "iphone", 4010.0: "ipad",
                     4039.0: "watch", 4069.0: "airpods",
                     4022.0: "iphone", 4012.0: "iphone"}


BUNDLE_CERT_BRANDS = {881.0, 885.0, 886.0, 888.0}   # 認證機/整新機品牌：bypass C3，DSS 一樣計主機


def _bundle_mach_cat(c3, c4, c6, brand=None):
    if c3 == 3001.0 or (brand in BUNDLE_CERT_BRANDS):
        if c4 in (4001.0, 4002.0):
            return "cpu"
        if c4 == 4004.0:
            return "iphone"
        if c4 in (4005.0, 4006.0, 4041.0):
            return "ipad"
        if c4 == 4038.0:
            return "watch"
    if c4 == 4014.0 and c6 in engine.C6_AIRPODS:
        return "airpods"
    return None


def _epb_bundle_query(shop_id, start_d, end_d):
    headers, raw = run_remote(f"""
select doc_id, trans_type, emp_id1, stk_qty,
       cat3_id, cat4_id, cat6_id, brand_id, coalesce(vip_id,'') vip,
       to_char(doc_date,'yyyymmdd') dd
from poslinev_bi
where shop_id = {quote_sql(str(shop_id))} and org_id = '01'
  and doc_date >= to_date('{start_d.isoformat()}','yyyy-mm-dd')
  and doc_date <  to_date('{(end_d + timedelta(days=1)).isoformat()}','yyyy-mm-dd')
order by doc_id, line_no
""", max_rows=200000)
    return headers, raw


def epb_bundle_stats(shop_id, start_d, end_d, _prefetched=None):
    """以 EPB poslinev_bi 重現 DSS 搭售統計。回傳格式同 dss_client.fetch_bundle_stats。

    _prefetched=(headers, raw)：跨期間共用同一次大區間查詢（VIP 配對以「日」為界，切片安全）。
    """
    headers, raw = _prefetched if _prefetched else _epb_bundle_query(shop_id, start_d, end_d)
    idx = {c: i for i, c in enumerate(headers)}
    lo, hi = start_d.strftime("%Y%m%d"), end_d.strftime("%Y%m%d")
    raw = [r for r in raw if lo <= r[idx["DD"]] <= hi]

    def numf(v):
        try:
            return float(str(v).replace(",", ""))
        except (TypeError, ValueError):
            return 0.0

    docs = {}
    for r in raw:
        # 僅 A=銷售 / E=銷退 / H=尾款 進統計（DSS 明細不含 G 訂金 / J 退訂）
        if r[idx["TRANS_TYPE"]] not in ("A", "E", "H"):
            continue
        docs.setdefault(r[idx["DOC_ID"]], []).append(r)

    def doc_vip(lines):
        v = (lines[0][idx["VIP"]] or "").strip()
        return v if v not in ("", "0") else None   # 無會員不做跨單配對

    # VIP-day 主機類別（3001 主機與 AirPods 分開記，歸類時 3001 優先）
    vipday_m3001, vipday_mair = {}, {}
    for lines in docs.values():
        vip = doc_vip(lines)
        if not vip:
            continue
        for r in lines:
            g = _bundle_mach_cat(numf(r[idx["CAT3_ID"]]), numf(r[idx["CAT4_ID"]]),
                                 numf(r[idx["CAT6_ID"]]), numf(r[idx["BRAND_ID"]]))
            key = (vip, r[idx["DD"]])
            if g == "airpods":
                vipday_mair.setdefault(key, set()).add(g)
            elif g:
                vipday_m3001.setdefault(key, set()).add(g)

    def pick_cat(own, cats):
        """own C4 類別優先；否則唯一主機類別；多類別且 own 不在其中 → own（可能 None）。"""
        if own in cats:
            return own
        if len(cats) == 1:
            return next(iter(cats))
        return own

    entries = []   # (emp, cat, kind B/A/Z, qty, vip, dd)
    free_acc = {}  # (vip, dd, cat) → 純配件單配件數
    for lines in docs.values():
        vip, dd = doc_vip(lines), lines[0][idx["DD"]]
        machs, accs = [], []
        for r in lines:
            c3, c4, c6 = numf(r[idx["CAT3_ID"]]), numf(r[idx["CAT4_ID"]]), numf(r[idx["CAT6_ID"]])
            g = _bundle_mach_cat(c3, c4, c6, numf(r[idx["BRAND_ID"]]))
            if g:
                machs.append((r, g))
            elif c3 == 3003.0:
                accs.append(r)
        m3001 = {g for _, g in machs if g != "airpods"}
        mair = {g for _, g in machs if g == "airpods"}
        acc_cat_count = {}
        for r in accs:
            own = BUNDLE_EPB_OWN_C4.get(numf(r[idx["CAT4_ID"]]))
            g, is_free = None, False
            # 歸類優先序：同單 3001 主機 → 同單 AirPods → 同 VIP 當日 3001 → 當日 AirPods
            for cats, free in ((m3001, False), (mair, False),
                               (vipday_m3001.get((vip, dd), set()) if vip else set(), True),
                               (vipday_mair.get((vip, dd), set()) if vip else set(), True)):
                if cats:
                    g, is_free = pick_cat(own, cats), free
                    break
            if g is None:
                continue   # 無任何主機脈絡 → 不列入統計（與 DSS 明細一致）
            if is_free:
                key = (vip, dd, g)
                free_acc[key] = free_acc.get(key, 0) + 1
            acc_cat_count[g] = acc_cat_count.get(g, 0) + 1
            entries.append((r[idx["EMP_ID1"]], g, "A", numf(r[idx["STK_QTY"]]), vip, dd))
        bycat = {}
        for r, g in machs:
            bycat.setdefault(g, []).append(r)
        for g, ms in bycat.items():
            for i, r in enumerate(ms):
                kind = "B" if (acc_cat_count.get(g, 0) > 0 and i == 0) else "Z"
                entries.append((r[idx["EMP_ID1"]], g, kind, numf(r[idx["STK_QTY"]]), vip, dd))

    has_b = {(vip, dd, g) for _, g, kind, _, vip, dd in entries if kind == "B" and vip}
    upgraded = set()
    fixed = []
    for emp, g, kind, qty, vip, dd in entries:
        key = (vip, dd, g)
        if (kind == "Z" and vip and free_acc.get(key, 0) > 0
                and key not in has_b and key not in upgraded):
            upgraded.add(key)
            kind = "B"
        fixed.append((emp, g, kind, qty))

    by_emp = {}
    for emp, g, kind, qty in fixed:
        slot = by_emp.setdefault(emp, {k: {"m": 0, "s": 0, "ms": 0} for k in
                                       ("cpu", "iphone", "ipad", "watch", "airpods")})
        field = {"B": "m", "A": "s", "Z": "ms"}[kind]
        slot[g][field] += int(qty)
    return [{"empId": emp, "empName": emp, **stats} for emp, stats in sorted(by_emp.items())]


# ─── 免登入門市清單 ────────────────────────────────────────────────────
def list_stores():
    """回傳所有可選門市。先嘗試 EPB 門市主檔（免使用者權限），
    失敗則退回後備硬編清單。回傳 {items:[{storeId,name}], default}。
    """
    items = []
    sql = (
        f"select shop_id, name from pos_shop "
        f"where org_id = {quote_sql(ORG_ID)} order by shop_id"
    )
    by_id = {}
    try:
        headers, rows = run_remote(sql, timeout=60)
        for row in rows:
            row = (row + [""] * len(headers))[: len(headers)]
            rec = {headers[i].upper(): row[i] for i in range(len(headers))}
            sid = text(rec.get("SHOP_ID"))
            if sid:
                by_id[sid] = text(rec.get("NAME")) or sid
    except Exception:
        by_id = {}
    # 補上後備清單中 EPB 未回傳的門市（如新門市 068），EPB 名稱優先
    for sid, name in FALLBACK_STORES.items():
        by_id.setdefault(sid, name)
    items = [{"storeId": sid, "name": by_id[sid]} for sid in sorted(by_id)]
    return {"items": items, "default": DEFAULT_SHOP_ID}


# ─── 員工解析 + 個人 sheet 動態重建（移植自 live-report-app）─────────────
def employees_from_sales(df, start_date, end_date, template_employees):
    if df.empty or "員工代碼" not in df.columns:
        return []
    template_names = {code: name for code, name in template_employees}
    d = engine.period(df, start_date, end_date)
    if d.empty:
        return []
    active = d[
        (d["員工代碼"].astype(str).str.strip() != "")
        & ((d["NET"].fillna(0).abs() > 0) | (d["數量"].fillna(0).abs() > 0))
    ].copy()
    if active.empty:
        return []
    employees = {}
    for _, row in active.iterrows():
        code = text(row.get("員工代碼"))
        if not code:
            continue
        name = text(row.get("員工名稱"))
        if not name or name == code:
            name = template_names.get(code, code)
        employees[code] = f"{code} {name}" if name and name != code else code
    return [(code, employees[code]) for code in sorted(employees)]


def resolve_report_employees(df_cur, dates, template_employees):
    employees = employees_from_sales(df_cur, dates["wk_start"], dates["wk_end"], template_employees)
    return employees or []


def copy_row_style(ws, source_row, target_row, max_col):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def rebuild_employee_sheet(ws, employees, start_row, template_employee_count, total_label="加總"):
    max_col = ws.max_column
    existing_total_row = start_row + template_employee_count
    data_style_row = start_row
    total_style_row = existing_total_row if existing_total_row <= ws.max_row else ws.max_row
    target_total_row = start_row + len(employees)
    clear_to = max(ws.max_row, target_total_row)

    for row in range(start_row, clear_to + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None
        copy_row_style(ws, data_style_row, row, max_col)

    for idx, (_, name) in enumerate(employees):
        row = start_row + idx
        ws.cell(row=row, column=1).value = name

    copy_row_style(ws, total_style_row, target_total_row, max_col)
    ws.cell(row=target_total_row, column=1).value = total_label
    if ws.max_row > target_total_row:
        ws.delete_rows(target_total_row + 1, ws.max_row - target_total_row)


def rebuild_employee_report_sheets(wb, employees, template_employee_count):
    rebuild_employee_sheet(wb["8.個人新制獎金"], employees, start_row=2, template_employee_count=template_employee_count)
    rebuild_employee_sheet(wb["9.個人週主機"], employees, start_row=3, template_employee_count=template_employee_count)
    rebuild_employee_sheet(wb["10.個人月主機"], employees, start_row=3, template_employee_count=template_employee_count)
    rebuild_employee_sheet(wb["11.個人月3PP"], employees, start_row=2, template_employee_count=template_employee_count)


def cell_has_sales_number(value):
    return isinstance(value, (int, float)) and abs(value) > 0.000001


def row_has_report_numbers(ws, row):
    for col in range(2, ws.max_column + 1):
        if cell_has_sales_number(ws.cell(row=row, column=col).value):
            return True
    return False


def employee_has_report_numbers(wb, employee_index):
    checks = [
        ("8.個人新制獎金", 2),
        ("9.個人週主機", 3),
        ("10.個人月主機", 3),
        ("11.個人月3PP", 2),
    ]
    for sheet_name, start_row in checks:
        if row_has_report_numbers(wb[sheet_name], start_row + employee_index):
            return True
    return False


def filter_employees_by_report_numbers(wb, employees):
    return [
        employee
        for index, employee in enumerate(employees)
        if employee_has_report_numbers(wb, index)
    ]


def finalize_known_formulas(wb):
    ws = wb["8.個人新制獎金"]
    for row in range(2, ws.max_row + 1):
        b = ws.cell(row=row, column=2).value or 0
        c = ws.cell(row=row, column=3).value or 0
        e = ws.cell(row=row, column=5).value or 0
        h = ws.cell(row=row, column=8).value or 0
        i = ws.cell(row=row, column=9).value or 0
        ws.cell(row=row, column=6).value = engine.safe_rate(e, e + b) if (e + b) else None
        ws.cell(row=row, column=7).value = engine.safe_rate(c, b) if b else None
        ws.cell(row=row, column=10).value = round(e / 2 / 1.05, 6) if e else None
        ws.cell(row=row, column=11).value = (h or 0) + (i or 0) + (ws.cell(row=row, column=10).value or 0)


def fill_employee_report_sheets(wb, df_cur, sacare_prices, dates):
    engine.fill_sheet6(wb["8.個人新制獎金"], df_cur, sacare_prices, dates)
    engine.fill_sheet78(wb["9.個人週主機"], wb["10.個人月主機"], df_cur, sacare_prices, dates)
    engine.fill_sheet9(wb["11.個人月3PP"], df_cur, sacare_prices, dates)
    finalize_known_formulas(wb)


def serialize_periods(dates, quarter_start, fiscal_start, q_num, w_num):
    payload = {key: value.isoformat() for key, value in dates.items()}
    payload["quarterStart"] = quarter_start.isoformat()
    payload["fiscalYearStart"] = fiscal_start.isoformat() if fiscal_start else ""
    payload["fiscalQuarter"] = q_num
    payload["fiscalWeek"] = w_num
    return payload


# ─── 報表組裝（移植自 live-report-app，加入 log 進度回呼）────────────────
def build_report_workbook(payload, log=lambda m: None):
    shop_id = safe_shop_id(payload.get("shopId"))
    overrides = payload.get("periods") or {}
    default_end = last_saturday()
    default_start = default_end - timedelta(days=6)
    # 本週基準：進階面板的 wkStart/wkEnd 優先，否則沿用主欄位
    wk_start = parse_date(overrides.get("wkStart") or payload.get("weekStart"), default_start)
    wk_end = parse_date(overrides.get("wkEnd") or payload.get("weekEnd"), default_end)
    if wk_end < wk_start:
        raise ValueError("本週結束日不可早於起始日")
    log(f"門市 {shop_id}　本週 {wk_start} ~ {wk_end}")
    dates = compute_periods(wk_start, wk_end)
    apply_period_overrides(dates, overrides, log=log)
    # 年對年截止日（Sheet 12/13）：前端可自訂，留空則沿用本週結束日
    yoy_end = parse_date(payload.get("yoyEnd"), wk_end)
    dates["yoy_end"] = yoy_end
    if yoy_end != wk_end:
        log(f"年對年截止日（自訂）：{yoy_end}（對比 {yoy_end.year-1} 年同期）")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_employees = engine.load_employees(wb) or engine.EMPLOYEES
    template_employee_count = len(template_employees)
    fiscal_start = engine.load_fiscal_year_start(wb)
    quarter_start = engine.compute_quarter_start(wk_start, fiscal_start)

    log("查詢 EPB 銷售資料…")
    df_cur, df_prev, sacare_prices, source_meta = load_epb_data(shop_id, dates, quarter_start)
    log(f"  取得 本期 {source_meta['currentRows']:,} 筆 / 去年 {source_meta['previousRows']:,} 筆")

    emp_count = _emp_count(payload)
    if emp_count:
        log(f"編制人數：{emp_count} 人（人均產值 = 營業額 / 編制人數）")
    traffic2, traffic10 = load_traffic(shop_id, dates, log=log)

    log("解析員工清單、重建個人 sheet…")
    employees = resolve_report_employees(df_cur, dates, template_employees)
    engine.EMPLOYEES = employees
    engine.EMP_CODES = [code for code, _ in employees]
    rebuild_employee_report_sheets(wb, employees, template_employee_count)
    log(f"  本週有交易員工 {len(employees)} 人")

    log("填入 1-5 報表…")
    engine.fill_sheet1(wb["1.主機銷售台數"], df_cur, quarter_start, wk_end)
    engine.fill_sheet2(wb["2.門市週報 "], df_cur, df_prev, sacare_prices, dates,
                       traffic=traffic2, emp_count=emp_count)
    engine.fill_sheet3(wb["3.3PP配件比較"], df_cur, df_prev, sacare_prices, dates)
    engine.fill_sheet45(wb["4.3PP 銷售排名"], wb["5.VAP銷售排名"], df_cur, sacare_prices, dates)

    log("填入 12-13 年對年報表…")
    engine.fill_sheet10(wb["12.月報YOY"], df_cur, df_prev, sacare_prices, dates,
                        traffic=traffic10, emp_count=emp_count)
    engine.fill_sheet11(wb["13.3PP YOY"], df_cur, df_prev, sacare_prices, dates)

    log("填入個人 8-11 報表…")
    fill_employee_report_sheets(wb, df_cur, sacare_prices, dates)

    active_employees = filter_employees_by_report_numbers(wb, employees)
    if active_employees != employees:
        log(f"  過濾無資料員工 → 保留 {len(active_employees)} 人，重算個人 sheet")
        engine.EMPLOYEES = active_employees
        engine.EMP_CODES = [code for code, _ in active_employees]
        rebuild_employee_report_sheets(wb, active_employees, len(employees))
        fill_employee_report_sheets(wb, df_cur, sacare_prices, dates)
        employees = active_employees

    bundle = load_bundle_stats(shop_id, wk_start, wk_end, fiscal_start, log=log)
    if bundle:
        log("填入 6-7 搭售統計…")
        fill_bundle_sheets(wb, bundle, employees, log=log)

    source_meta["employeeCount"] = len(employees)
    source_meta["employees"] = [{"code": code, "name": name} for code, name in employees]

    # 本 app 輸出不含「設定」工作表（設定僅為共用模板的設定來源，仍保留在模板供讀取）
    if "設定" in wb.sheetnames:
        del wb["設定"]

    # 修正 9/10 頁：引擎 fill_sheet78 從第 3 列起寫資料，第 2 列恆為空白 → 移除該空列
    for sheet_name in ("9.個人週主機", "10.個人月主機"):
        ws_fix = wb[sheet_name]
        if all(ws_fix.cell(row=2, column=c).value in (None, "") for c in range(1, ws_fix.max_column + 1)):
            ws_fix.delete_rows(2, 1)

    q_num, w_num = engine.compute_fiscal_week_number(wk_start, fiscal_start) if fiscal_start else ("", "")
    return wb, {
        "ok": True,
        "source": "epb",
        "sourceMeta": source_meta,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "periods": serialize_periods(dates, quarter_start, fiscal_start, q_num, w_num),
    }


def build_report_excel(payload, log=lambda m: None):
    wb, meta = build_report_workbook(payload, log=log)
    log("儲存 Excel…")
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    periods = meta["periods"]
    shop_id = meta["sourceMeta"].get("shopId", safe_shop_id(payload.get("shopId")))
    filename = f"門市報表_{shop_id}_{periods['wk_start']}~{periods['wk_end']}.xlsx"
    return filename, stream.getvalue(), meta


# ─── Job 管理（仿北一區週報-app）───────────────────────────────────────
def _run_job(job_id, payload):
    def log(msg):
        ts = time.strftime("%H:%M:%S")
        with _LOCK:
            JOBS[job_id]["messages"].append(f"[{ts}] {msg}")

    with _LOCK:
        JOBS[job_id]["status"] = "running"

    try:
        wk_end = date.fromisoformat(payload["weekEnd"])
        payload.setdefault("weekStart", (wk_end - timedelta(days=6)).isoformat())
        # 記住本店編制人數，下次預設帶入
        if traffic_mod is not None and payload.get("employeeCount") not in (None, ""):
            try:
                traffic_mod.set_employee_count(payload["shopId"], payload["employeeCount"])
            except Exception:
                pass
        filename, body, meta = build_report_excel(payload, log=log)
        with _LOCK:
            JOBS[job_id]["status"] = "done"
            JOBS[job_id]["result"] = body
            JOBS[job_id]["filename"] = filename
            JOBS[job_id]["meta"] = meta
        log(f"✓ 完成！{filename}（{len(body):,} bytes）")
    except Exception:
        tb = traceback.format_exc()
        with _LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["error"] = tb
        log(f"✗ 錯誤:\n{tb}")


# ─── HTTP Handler ─────────────────────────────────────────────────────
class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_ROOT), **kwargs)

    def log_message(self, fmt, *args):
        pass

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False, default=json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_xlsx(self, filename, body):
        quoted = urllib.parse.quote(filename)
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quoted}")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_png(self, body):
        self.send_response(200)
        self.send_header("Content-Type", "image/png")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        n = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(n).decode("utf-8") if n else "{}")

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        qs = urllib.parse.parse_qs(p.query)
        try:
            if p.path == "/api/default-date":
                self.send_json(200, {"date": last_saturday().isoformat()})
                return
            if p.path == "/api/stores":
                self.send_json(200, list_stores())
                return
            if p.path == "/api/periods":
                wk_end_s = qs.get("weekEnd", [""])[0]
                if not wk_end_s:
                    self.send_json(400, {"error": "缺少 weekEnd"})
                    return
                wk_end = parse_date(wk_end_s)
                wk_start = parse_date(qs.get("wkStart", [""])[0], wk_end - timedelta(days=6))
                if qs.get("wkEnd", [""])[0]:
                    wk_end = parse_date(qs.get("wkEnd", [""])[0])
                d = compute_periods(wk_start, wk_end)
                self.send_json(200, {
                    "prevWkStart": d["prev_wk_start"].isoformat(),
                    "prevWkEnd": d["prev_wk_end"].isoformat(),
                    "wkStart": d["wk_start"].isoformat(),
                    "wkEnd": d["wk_end"].isoformat(),
                    "moStart": d["mo_start"].isoformat(),
                    "moEnd": d["mo_end"].isoformat(),
                    "lmStart": d["lm_start"].isoformat(),
                    "lmEnd": d["lm_end"].isoformat(),
                    "lyStart": d["ly_start"].isoformat(),
                    "lyEnd": d["ly_end"].isoformat(),
                })
                return
            if p.path == "/api/shoppertrak/status":
                if traffic_mod is None:
                    self.send_json(200, {"available": False, "hasCredentials": False,
                                         "username": ""})
                    return
                u, _ = traffic_mod.get_credentials()
                self.send_json(200, {
                    "available": True,
                    "hasCredentials": traffic_mod.has_credentials(),
                    "username": u or "",
                })
                return
            if p.path == "/api/bundle/preview":
                # 搭售統計兩源比對：?shopId=004&start=YYYY-MM-DD&end=YYYY-MM-DD&source=epb|dss|both
                shop_id = safe_shop_id(qs.get("shopId", [""])[0] or DEFAULT_SHOP_ID)
                start_s, end_s = qs.get("start", [""])[0], qs.get("end", [""])[0]
                if not (start_s and end_s):
                    self.send_json(400, {"error": "缺少 start / end（YYYY-MM-DD）"})
                    return
                start_d, end_d = parse_date(start_s), parse_date(end_s)
                source = (qs.get("source", ["both"])[0] or "both").lower()
                resp = {"shopId": shop_id, "start": start_d.isoformat(), "end": end_d.isoformat()}
                if source in ("epb", "both"):
                    resp["epb"] = epb_bundle_stats(shop_id, start_d, end_d)
                if source in ("dss", "both"):
                    try:
                        resp["dss"] = dss_mod.fetch_bundle_stats(shop_id, start_d, end_d) if dss_mod else None
                    except Exception as exc:
                        resp["dssError"] = str(exc)
                self.send_json(200, resp)
                return
            if p.path == "/api/dss/status":
                force = bool(traffic_mod and traffic_mod.load_config().get("bundleSource") == "dss")
                if dss_mod is None:
                    self.send_json(200, {"available": False, "hasCredentials": False,
                                         "username": "", "state": "idle", "error": "",
                                         "force": force})
                    return
                payload = dss_mod.status()
                payload["force"] = force
                self.send_json(200, payload)
                return
            if p.path == "/api/dss/captcha":
                png = dss_mod.get_captcha_png() if dss_mod else None
                if not png:
                    self.send_json(404, {"error": "目前沒有驗證碼，請先按「登入 DSS」"})
                    return
                self.send_png(png)
                return
            if p.path == "/api/config":
                shop_id = safe_shop_id(qs.get("shopId", [""])[0] or DEFAULT_SHOP_ID)
                emp = traffic_mod.get_employee_count(shop_id) if traffic_mod else None
                has_site = bool(traffic_mod and traffic_mod.site_id_for_shop(shop_id))
                self.send_json(200, {"employeeCount": emp, "hasSiteId": has_site})
                return
            if p.path == "/api/status":
                job_id = qs.get("jobId", [""])[0]
                with _LOCK:
                    job = JOBS.get(job_id)
                if not job:
                    self.send_json(404, {"error": "找不到工作"})
                    return
                resp = {
                    "status": job["status"],
                    "messages": list(job["messages"]),
                    "filename": job.get("filename", ""),
                }
                if job["status"] == "error":
                    resp["error"] = job.get("error", "未知錯誤")
                self.send_json(200, resp)
                return
            if p.path == "/api/download":
                job_id = qs.get("jobId", [""])[0]
                with _LOCK:
                    job = JOBS.get(job_id)
                if not job or job["status"] != "done":
                    self.send_json(400, {"error": "檔案尚未就緒"})
                    return
                self.send_xlsx(job["filename"], job["result"])
                return
        except Exception as exc:
            self.send_json(500, {"error": str(exc)})
            return
        super().do_GET()

    def do_POST(self):
        p = urllib.parse.urlparse(self.path)
        if p.path == "/api/generate":
            try:
                payload = self.read_body()
                shop_id = safe_shop_id(payload.get("shopId"))
                wk_end_s = str(payload.get("weekEnd", payload.get("week_end", ""))).strip()
                date.fromisoformat(wk_end_s)
                payload["shopId"] = shop_id
                payload["weekEnd"] = wk_end_s
            except Exception as e:
                self.send_json(400, {"error": f"參數錯誤: {e}"})
                return
            job_id = str(uuid.uuid4())
            now = time.time()
            with _LOCK:
                # 順手清掉超過 24 小時的舊工作，避免長期運行記憶體無限增長
                for old_id in [k for k, v in JOBS.items() if now - v.get("created", now) > 86400]:
                    del JOBS[old_id]
                JOBS[job_id] = {"status": "pending", "messages": [], "result": None, "created": now}
            threading.Thread(
                target=_run_job, args=(job_id, payload), daemon=True
            ).start()
            self.send_json(200, {"jobId": job_id})
            return
        if p.path.startswith("/api/dss/"):
            if dss_mod is None:
                self.send_json(400, {"error": "DSS 模組未載入"})
                return
            try:
                body = self.read_body()
            except Exception as e:
                self.send_json(400, {"error": f"參數錯誤: {e}"})
                return
            if p.path == "/api/dss/credentials":
                if body.get("clear"):
                    dss_mod.clear_credentials()
                    self.send_json(200, {"ok": True, "hasCredentials": False})
                    return
                username = str(body.get("username", "")).strip()
                password = str(body.get("password", ""))
                if not username or not password:
                    self.send_json(400, {"error": "請輸入帳號與密碼"})
                    return
                dss_mod.set_credentials(username, password)
                self.send_json(200, {"ok": True, "hasCredentials": True, "username": username})
                return
            if p.path == "/api/dss/force":
                # 勾選「強制使用 DSS」：bundleSource=dss（不退回 EPB）；取消 → 預設 EPB 為主
                if traffic_mod is None:
                    self.send_json(400, {"error": "設定模組未載入"})
                    return
                cfg = traffic_mod.load_config()
                if body.get("force"):
                    cfg["bundleSource"] = "dss"
                else:
                    cfg.pop("bundleSource", None)
                traffic_mod.save_config(cfg)
                self.send_json(200, {"ok": True, "force": bool(body.get("force"))})
                return
            if p.path == "/api/dss/login/start":
                self.send_json(200, dss_mod.start_login())
                return
            if p.path == "/api/dss/login/refresh-captcha":
                self.send_json(200, dss_mod.refresh_captcha())
                return
            if p.path == "/api/dss/login/captcha":
                self.send_json(200, dss_mod.submit_captcha(str(body.get("code", ""))))
                return
            if p.path == "/api/dss/login/otp":
                self.send_json(200, dss_mod.submit_otp(str(body.get("code", ""))))
                return
            self.send_json(404, {"error": "Not found"})
            return
        if p.path == "/api/shoppertrak/credentials":
            if traffic_mod is None:
                self.send_json(400, {"error": "來客數模組未載入"})
                return
            try:
                body = self.read_body()
            except Exception as e:
                self.send_json(400, {"error": f"參數錯誤: {e}"})
                return
            if body.get("clear"):
                traffic_mod.clear_credentials()
                self.send_json(200, {"ok": True, "hasCredentials": False})
                return
            username = str(body.get("username", "")).strip()
            password = str(body.get("password", ""))
            if not username or not password:
                self.send_json(400, {"error": "請輸入帳號與密碼"})
                return
            traffic_mod.set_credentials(username, password)
            self.send_json(200, {"ok": True, "hasCredentials": True, "username": username})
            return
        self.send_json(404, {"error": "Not found"})


def main():
    for label, path in (("java", JAVA), ("javac", JAVAC)):
        if not Path(path).exists():
            sys.exit(
                f"找不到 {label}：{path}\n"
                "請安裝 JDK 1.8，或設定環境變數 EPB_JAVA_HOME 指到 JDK Home 目錄，例如：\n"
                "  export EPB_JAVA_HOME=/Library/Java/JavaVirtualMachines/jdk1.8.0_251.jdk/Contents/Home"
            )
    compile_java("EPBReportQuery.java")
    if dss_mod is not None:
        # 嘗試還原上次 DSS session（背景執行，不阻擋啟動）
        def _dss_restore():
            try:
                if dss_mod.restore_session():
                    print("DSS：已還原上次登入 session", flush=True)
            except Exception:
                pass
        threading.Thread(target=_dss_restore, daemon=True).start()
    port = int(os.environ.get("PORT", "8783"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"門市報表製作：http://127.0.0.1:{port}", flush=True)
    print("按 Ctrl+C 停止", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("已停止。")


if __name__ == "__main__":
    main()
