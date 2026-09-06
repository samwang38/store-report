#!/usr/bin/env python3
"""
週報 Excel 自動填充腳本
Usage:
  python fill_weekly_excel.py \
    --week-start 2026-04-05 \
    --week-end   2026-04-11 \
    --data-dir   /path/to/銷售資料/ \
    --template   /path/to/週報模板.xlsx \
    --output     /path/to/週報_2026-04-05~04-11.xlsx
"""
from __future__ import annotations

VERSION = "1.2.0"

import argparse, glob, shutil, sys
from datetime import date, timedelta
from calendar import monthrange
from pathlib import Path

import pandas as pd
import openpyxl

# ─── Employee list ─────────────────────────────────────────────────────────────
# Default list (used as fallback if template has no 設定 sheet).
# At runtime, main() overrides these from the template's 設定 sheet,
# so other branches only need to edit that sheet — no Python changes required.
EMPLOYEES = [
    ('SA1092', 'Sam'), ('SA1614', 'Dean'), ('SA1893', 'Jun'),
    ('SA2765', 'Nana'), ('SA3011', 'Zoe'), ('SA3154', 'Winter'), ('SA3221', 'Tanya'),
]
EMP_CODES = [e[0] for e in EMPLOYEES]

def load_employees(wb):
    """Read employee list from 設定 sheet.
    Columns: A = 員工代碼, B = 顯示名稱, C = 排除毛利 (已廢棄，忽略), D = 年度起始日.
    Returns list of (code, name) tuples, or None if sheet is absent / empty."""
    if '設定' not in wb.sheetnames:
        return None
    ws = wb['設定']
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0]).strip() if row[0] is not None else ''
        name = str(row[1]).strip() if row[1] is not None else ''
        if code and name:
            result.append((code, name))
    return result or None

def load_fiscal_year_start(wb) -> 'date | None':
    """Read fiscal year start date from 設定 sheet col D row 2.
    Returns a date object, or None if not set (will fall back to calendar-quarter logic)."""
    if '設定' not in wb.sheetnames:
        return None
    ws = wb['設定']
    val = ws.cell(row=2, column=4).value   # D2
    if val is None:
        return None
    if hasattr(val, 'date'):               # datetime → date
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip())
    except ValueError:
        return None

# ─── C6 device codes (Mac only — C6 still needed to split mini/iMac/MacBook) ───
C6_CPU     = {6001.0, 6002.0, 6007.0, 6008.0, 6340.0, 6341.0, 6342.0, 6343.0, 6344.0}
#  6001 = iMac, 6002 = Mac mini
#  6340 = MBA 13 M5, 6341 = MBA 15 M5, 6342 = MBP 14 M4/M5, 6343 = MBP 16 M5, 6344 = MacBook Neo
#  ↑ 新 Mac 世代上市時只需更新此處，其餘計算自動跟進
C6_MACBOOK = C6_CPU - {6001.0, 6002.0}   # MacBook 系列（排除 mini / iMac）
C6_AIRPODS = {6258.0, 6312.0, 6330.0}
C6_SA      = {'cpu': {6533.0}, 'ipad': {6534.0}, 'iphone': {6535.0},
              'watch': {6536.0}, 'airpods': {6537.0}}

# ─── C4 device codes (stable across model years — use for iPhone/Watch/iPad) ───
# C4 represents product category and does NOT change when new models are released.
# Confirmed consistent between 2025 (iPhone 16 era) and 2026 (iPhone 17 era).
C4_IPHONE  = {4004.0}                    # all iPhone models, all years
C4_IPAD    = {4005.0, 4006.0, 4041.0}   # iPad std/Air(4005) + mini(4006) + Pro(4041)
C4_WATCH   = {4038.0}                    # all Watch models, all years
# Mac: C4=4002→MacBook, C4=4001→Mac mini+iMac (split still done via C6)

# C4 categories for 3PP accessory sheet
C4_ROWS = {
    4007.0: 2, 4009.0: 3, 4010.0: 4, 4012.0: 5, 4013.0: 6, 4014.0: 7,
    4017.0: 8, 4021.0: 9, 4022.0: 10, 4026.0: 11, 4039.0: 12,
    4050.0: 13, 4053.0: 14, 4069.0: 15,
}
C4_CPU_SET     = {4007.0}
C4_IPHONE_SET  = {4009.0}
C4_IPAD_SET    = {4010.0}
C4_WATCH_SET   = {4039.0}
C4_AIRPODS_SET = {4069.0}
VAP_BRANDS     = {59.0, 224.0, 277.0, 106.0}

# ─── Sheet 3/4（每月重點 / Speakers）───────────────────────────────────────────
# 每月重點：本期統計對象為品牌代碼 496（Starter Kit / Hello Nature 等入門組）。
# 此頁每月會換統計對象、且未必以品牌區分，換檔時直接改這裡。
MONTHLY_FOCUS_BRANDS = {496.0}
SPEAKERS_C4          = {4013.0}   # 3PP 藍牙喇叭（同第 5 頁 3PP配件比較的 Speakers 列）

# ─── D 欄「原廠商品營業額」排除條件 ──────────────────────────────────────────
C1_EXCLUDED_FROM_APPLE = {1002.0, 1004.0, 1008.0}
C3_EXCLUDED_FROM_APPLE = {3047.0, 3003.0, 3004.0, 3018.0, 3019.0, 3012.0}
C6_EXCLUDED_FROM_APPLE = {6888.0, 6889.0}
SKU_EXCLUDED_FROM_APPLE = {
    '99200168', '99500006', '99900946', '99900947', '99900948', '99900949',
    '99900950', '99901684', '99901685', '99902607', '99902608', '99902609',
    '99902610',
    '888',       # 2點折1元（點數折抵）
    '99200202',  # 教育價活動（教育優惠內部代碼）
    '99901780',  # SA Care 檢測新機活動代碼
}

# Slide 1 product row definitions
PRODUCT_ROWS = {
    1:  (3001.0, {6001.0}, None, None),
    2:  (3001.0, {6002.0}, None, None),
    3:  (3001.0, None, 'MBP 14', None),   # 名稱辨別，自動涵蓋各世代 MacBook Pro 14
    4:  (3001.0, None, 'MBP 16', None),   # MacBook Pro 16
    5:  (3001.0, None, 'MBA 13', None),   # MacBook Air 13
    6:  (3001.0, None, 'MBA 15', None),   # MacBook Air 15
    7:  (3001.0, None, 'MBN',    None),   # MacBook Neo
    # ── row 8 = CPU 小計（由 SUBTOTAL_ROWS 寫入）──
    9:  (3002.0, {6330.0}, None, None),
    10: (3002.0, {6312.0}, '主動式降噪', None),
    11: (3002.0, {6312.0}, None, '主動式降噪'),
    12: (3002.0, {6258.0}, None, None),
    13: (3002.0, {6073.0}, 'MINI', None),
    14: (3002.0, {6073.0}, None, 'MINI'),
    16: (3001.0, {6327.0}, None, None),
    17: (3001.0, {6328.0}, None, None),
    18: (3001.0, {6329.0}, None, None),
    20: (3001.0, {6323.0}, None, None),
    21: (3001.0, {6324.0}, None, None),
    22: (3001.0, {6325.0}, None, None),
    23: (3001.0, {6326.0}, None, None),
    24: (3001.0, {6335.0}, None, None),
    26: (3001.0, {6321.0, 6322.0}, None, None),
    27: (3001.0, {6317.0, 6318.0, 6336.0}, None, None),
    28: (3001.0, {6319.0, 6338.0}, None, None),
    29: (3001.0, {6331.0, 6332.0}, None, None),
    30: (3001.0, {6333.0}, None, None),
    31: (3001.0, {6313.0}, None, None),
    33: (3002.0, {6100.0}, '第一代', None),
    34: (3002.0, {6100.0}, '2nd', None),
    35: (3002.0, {6100.0}, 'USB-C', None),
    36: (3002.0, {6100.0}, 'PRO', 'keyboard'),  # 排除 Magic Keyboard for iPad Pro
}
SUBTOTAL_ROWS = {
    8:  list(range(1, 8)),   # CPU 小計（含 MacBook Neo row 7）
    15: list(range(9, 15)),  # AirPods 小計
    19: list(range(16, 19)),
    25: list(range(20, 25)),
    32: list(range(26, 32)),
    37: list(range(33, 37)), # Apple Pencil 小計
}

# ─── CLI ───────────────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--version',    action='store_true', help='顯示版本號後結束')
    p.add_argument('--week-start', required=False)
    p.add_argument('--week-end',   required=False)
    p.add_argument('--yoy-end',    required=False, help='Sheet 12/13 年對年截止日（YYYY-MM-DD）；預設＝週結束日')
    p.add_argument('--data-dir',   required=False)
    p.add_argument('--template',   required=False)
    p.add_argument('--output',     required=False)
    p.add_argument('--output-dir', required=False)
    return p.parse_args()

# ─── File discovery ────────────────────────────────────────────────────────────
def find_800ab(data_dir: Path, year: int) -> str | None:
    # Try year-prefixed patterns first
    for pat in [f'800AB_{year}*.xlsx', f'800AB{year}*.xlsx']:
        found = sorted(glob.glob(str(data_dir / pat)))
        if found:
            if year < date.today().year:
                annual = [f for f in found if '整年' in f]
                return annual[0] if annual else found[-1]
            else:
                non_annual = [f for f in found if '整年' not in f]
                return non_annual[-1] if non_annual else found[-1]
    # Fallback: any 800AB*.xlsx without 整年 for current year,
    # or with 整年 for previous year
    all_files = sorted(glob.glob(str(data_dir / '800AB*.xlsx')))
    if year < date.today().year:
        annual = [f for f in all_files if '整年' in f and str(year) in f]
        return annual[0] if annual else None
    else:
        non_annual = [f for f in all_files if '整年' not in f]
        return non_annual[-1] if non_annual else None

def find_sacare(data_dir: Path) -> Path | None:
    for name in ['SAcare對應價目表.xlsx', 'SAcare對應價目表.xlsx']:
        p = data_dir / name
        if p.exists():
            return p
    return None

# ─── Data loading ──────────────────────────────────────────────────────────────
def find_header_row(filepath):
    for h in range(0, 20):
        try:
            df = pd.read_excel(filepath, header=h, nrows=1)
            # 去除 BOM 及前後空白再比對，避免欄位名稱有隱藏字元導致比對失敗
            cols = {str(c).strip().lstrip('\ufeff') for c in df.columns.tolist()}
            # 不限定欄位順序，只要 '單據日期' 與 '單據代碼' 同時出現即為正確 header 列
            if '單據日期' in cols and '單據代碼' in cols:
                return h
        except Exception:
            pass
    return 8  # 自店 004 預設 header 在第 9 行（0-indexed = 8）

# 所有報表計算所需的必要欄位，及其用途說明（用於錯誤訊息）
REQUIRED_COLUMNS = {
    '單據日期':     '期間篩選（本週 / 本月 / 去年同期）',
    '單據代碼':     '成交筆數計算',
    '交易類型':     '銷售=一般／訂金=定金計營業額／尾款=完成計台數／銷退=退貨扣除',
    '存貨代碼':     '辨識 SAcare 品項',
    '名稱':         'MacBook 機型區分、ACPP、Office、3PP 排名',
    '數量':         '台數與數量計算',
    '銷售金額(含稅)':'營業額計算',
    '銷退金額':     '營業額計算（退貨扣除）',
    '單位成本':     '毛利計算',
    '品牌代碼':     '認證機台數（881/885/886/888）、VAP 業績、毛利排除（297）',
    '類別1代碼':    '毛利排除條件（C1=21）',
    '類別3代碼':    '主要品類（3001 主機 / 3002 原廠配件 / 3003 三方 / 3032 ACPP）',
    '類別4代碼':    'iPhone / iPad / Watch 台數（C4 穩定代碼）',
    '類別6代碼':    'Mac 機型區分、SAcare 品類、Sheet 1 產品列',
    '員工代碼':         '個人業績計算（Sheet 6～9）',
    '折扣':             '排除贈品交易（Sheet 4/5 銷售排名）',
    '等級代碼':         'Sheet 6 排除等級 05（非全職／兼職員工）',
    '淨銷售金額(未稅)': 'Sheet 6 H/I 欄毛利計算主要依據（缺少時毛利會變負數）',
}

def validate_columns(df, filepath):
    """Check all required columns exist; exit with clear message if any are missing."""
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        print(f'\n❌ 錯誤：800AB 檔案缺少必要欄位')
        print(f'   檔案：{Path(filepath).name}')
        print(f'\n   缺少的欄位（共 {len(missing)} 個）：')
        for col in missing:
            print(f'     • {col}　→　{REQUIRED_COLUMNS[col]}')
        print(f'\n   請重新匯出 800AB，確認以上欄位已勾選後再執行。')
        sys.exit(1)

def load_800(filepath):
    print(f'  Loading {Path(filepath).name}...', flush=True)
    hdr = find_header_row(filepath)
    try:
        df = pd.read_excel(filepath, sheet_name='Sheet', header=hdr)
    except Exception:
        df = pd.read_excel(filepath, header=hdr)
    # 統一去除欄位名稱的 BOM 及前後空白，防止 KeyError
    df.columns = [str(c).strip().lstrip('\ufeff') for c in df.columns]
    df['單據日期'] = pd.to_datetime(df['單據日期'])
    for col in ['品牌代碼', '類別3代碼', '類別4代碼', '類別6代碼']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    if '員工代碼' in df.columns:
        df['員工代碼'] = df['員工代碼'].astype(str).str.strip()
    if '等級代碼' in df.columns:
        df['等級代碼'] = df['等級代碼'].astype(str).str.strip()
    df['NET'] = df.get('銷售金額(含稅)', 0).fillna(0) + df.get('銷退金額', 0).fillna(0)
    validate_columns(df, filepath)
    return df

def load_sacare(path: Path) -> dict:
    print(f'  Loading SAcare prices...', flush=True)
    dfs = pd.read_excel(path, header=0)
    prices = {}
    for _, row in dfs.iterrows():
        if pd.notna(row.iloc[0]) and pd.notna(row.iloc[2]):
            prices[str(row.iloc[0]).strip()] = float(row.iloc[2])
    print(f'    {len(prices)} SAcare codes', flush=True)
    return prices

# ─── Date utilities ────────────────────────────────────────────────────────────
def period(df, start: date, end: date) -> pd.DataFrame:
    s, e = pd.Timestamp(start), pd.Timestamp(f'{end} 23:59:59')
    return df[(df['單據日期'] >= s) & (df['單據日期'] <= e)]

def compute_quarter_start(week_start: date, fiscal_year_start: 'date | None' = None) -> date:
    """Return the start date of the fiscal quarter containing week_start.

    If fiscal_year_start is provided (read from 設定 sheet col D2):
      - Fiscal year = 52 weeks (4 × 13), origin = fiscal_year_start
      - Quarter start = fiscal_year_start + (quarter_index × 13 weeks)

    Falls back to calendar-quarter logic if fiscal_year_start is None.
    """
    if fiscal_year_start is not None:
        delta_days = (week_start - fiscal_year_start).days
        week_in_year = delta_days // 7          # 0-based week index from year start
        quarter_idx  = week_in_year // 13       # 0-based quarter index (0=Q1,1=Q2…)
        return fiscal_year_start + timedelta(weeks=quarter_idx * 13)

    # Fallback: calendar-quarter (original logic)
    q_first_month = ((week_start.month - 1) // 3) * 3 + 1
    q_first_day   = date(week_start.year, q_first_month, 1)
    days_ahead    = (week_start.weekday() - q_first_day.weekday()) % 7
    return q_first_day + timedelta(days=days_ahead)


def compute_fiscal_week_number(week_start: date, fiscal_year_start: date) -> tuple:
    """Returns (quarter_num, week_in_quarter) — both 1-based."""
    delta_days      = (week_start - fiscal_year_start).days
    week_in_year    = delta_days // 7
    quarter_num     = week_in_year // 13 + 1
    week_in_quarter = week_in_year % 13 + 1
    return quarter_num, week_in_quarter

def safe_rate(num, den):
    if not den or num is None:
        return 0
    return round(num / den, 4)

# 台數在尾款當天計入（銷售完成日）；訂金只計營業額，不計台數
# 一般現金交易用「銷售」；分期付款完成用「尾款」
SALE_TYPES = {'銷售', '尾款'}

# ─── Core metric calculator ────────────────────────────────────────────────────
def calc_metrics(d: pd.DataFrame, sacare_prices: dict) -> dict:
    """Calculate all KPIs for a filtered DataFrame (already period-filtered)."""
    sa_codes = set(sacare_prices.keys())
    non_sa = d[~d['存貨代碼'].astype(str).str.strip().isin(sa_codes)]

    # Revenue by category
    rev_3001 = non_sa.loc[non_sa['類別3代碼'] == 3001.0, 'NET'].sum()
    rev_3002 = non_sa.loc[non_sa['類別3代碼'] == 3002.0, 'NET'].sum()
    rev_3003 = non_sa.loc[non_sa['類別3代碼'] == 3003.0, 'NET'].sum()

    # SAcare：只計 SALE_TYPES（銷售/尾款），排除訂金避免雙重計算
    sa_rows = d[d['存貨代碼'].astype(str).str.strip().isin(sa_codes)].copy()
    sa_sold = sa_rows[sa_rows['交易類型'].isin(SALE_TYPES)]
    sa_ret  = sa_rows[sa_rows['交易類型'] == '銷退']
    sa_sold_net = (sa_sold['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa_sold['數量'].fillna(0)).sum()
    sa_ret_net  = (sa_ret['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa_ret['數量'].abs()).sum()
    sa_rev = sa_sold_net - sa_ret_net
    sa_gross = sa_rev / 2

    # 總營業額 = ALL non-SA NET (incl. voucher/misc deductions) + SA_rev
    total_rev = non_sa['NET'].sum() + sa_rev

    # Gross profit (Apple C3 ∈ {3001,3002,3032,3033}, excl SAcare, brand≠297, C6≠31, C1≠21)
    excl = (d['存貨代碼'].astype(str).str.strip().isin(sa_codes) |
            (d.get('品牌代碼', pd.Series(dtype=float)) == 297.0) |
            (d.get('類別6代碼', pd.Series(dtype=float)) == 31.0) |
            (d.get('類別1代碼', pd.Series(dtype=float)) == 21.0))
    apl_mask = d['類別3代碼'].isin([3001.0, 3002.0, 3032.0, 3033.0]) & ~excl
    tpp_mask = (d['類別3代碼'] == 3003.0) & ~excl

    def gross(mask):
        # 排除尾款：尾款 NET=0 但 ERP 仍記錄完整成本，會造成毛利負數
        sub = d[mask & (d['交易類型'] != '尾款')]
        return (sub['NET'] - (sub.get('單位成本', 0).fillna(0) * 1.05).round() * sub['數量'].fillna(0)).sum()

    apl_gross = gross(apl_mask)
    tpp_gross = gross(tpp_mask)

    # Unit counts
    # Mac: C6-based (need to distinguish MacBook / mini / iMac within same C4=4001/4002)
    cert_brands = {881.0, 885.0, 886.0, 888.0}
    unit_mask = (d['類別3代碼'] == 3001.0) | (d['品牌代碼'].isin(cert_brands))

    def net_units_c6(c6_set):
        m = unit_mask & d['類別6代碼'].isin(c6_set)
        sale = d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum()
        ret  = d.loc[m & (d['交易類型'] == '銷退'),  '數量'].abs().sum()
        return int(sale - ret)

    # iPhone / Watch / iPad: C4-based + 認證機品牌 bypass
    def net_units_c4(c4_set):
        m = ((d['類別3代碼'] == 3001.0) | d['品牌代碼'].isin(cert_brands)) & \
            d['類別4代碼'].isin(c4_set)
        sale = d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum()
        ret  = d.loc[m & (d['交易類型'] == '銷退'),  '數量'].abs().sum()
        return int(sale - ret)

    cpu_non_mini = net_units_c6(C6_MACBOOK)   # MacBook（引用 C6_MACBOOK 常數，新世代只需更新 C6_CPU）
    cpu_mini     = net_units_c6({6002.0})
    cpu_imac     = net_units_c6({6001.0})
    cpu_total    = cpu_non_mini + cpu_mini + cpu_imac
    iphone_units = net_units_c4(C4_IPHONE)
    ipad_units   = net_units_c4(C4_IPAD)
    watch_units  = net_units_c4(C4_WATCH)

    # ACPP-MAC
    acpp_mac = int(d.loc[(d['類別3代碼'] == 3032.0) &
                         d['名稱'].str.lower().str.contains('mac', na=False) &
                         d['交易類型'].isin(SALE_TYPES), '數量'].sum())

    # SAcare counts by device
    def sa_units(c6_set):
        m = d['類別6代碼'].isin(c6_set) & d['存貨代碼'].astype(str).str.strip().isin(sa_codes)
        return int(d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum() -
                   d.loc[m & (d['交易類型'] == '銷退'), '數量'].abs().sum())

    sa_cpu     = sa_units(C6_SA['cpu'])
    sa_ipad    = sa_units(C6_SA['ipad'])
    sa_iphone  = sa_units(C6_SA['iphone'])
    sa_watch   = sa_units(C6_SA['watch'])
    sa_airpods = sa_units(C6_SA['airpods'])

    # 成交筆數 (excluding 正負沖銷)
    sale_codes = set(d.loc[d['交易類型'] == '銷售', '單據代碼'].dropna().astype(str))
    ret_codes  = set(d.loc[d['交易類型'] == '銷退',  '單據代碼'].dropna().astype(str))
    txn_count  = len(sale_codes - ret_codes)

    return dict(
        rev_3001=int(rev_3001), rev_3002=int(rev_3002), rev_3003=int(rev_3003),
        sa_rev=int(sa_rev), sa_gross=int(sa_gross),
        total_rev=int(total_rev),
        apl_gross=int(apl_gross), tpp_gross=int(tpp_gross),
        total_gross=int(apl_gross + tpp_gross + sa_gross),
        cpu_non_mini=cpu_non_mini, cpu_mini=cpu_mini, cpu_total=cpu_total,
        iphone=iphone_units, ipad=ipad_units, watch=watch_units,
        acpp_mac=acpp_mac,
        sa_cpu=sa_cpu, sa_ipad=sa_ipad, sa_iphone=sa_iphone,
        sa_watch=sa_watch, sa_airpods=sa_airpods,
        txn_count=txn_count,
    )

# ─── Sheet 1: 主機銷售台數 ─────────────────────────────────────────────────────
def fill_sheet1(ws, df_cur: pd.DataFrame, quarter_start: date, week_end: date):
    print('  Sheet 1: 主機銷售台數', flush=True)
    d = df_cur.copy()   # 尾款計台數，訂金不計，透過 SALE_TYPES 控制
    _cert = {881.0, 885.0, 886.0, 888.0}
    _brand = d.get('品牌代碼', pd.Series(dtype=float, index=d.index)).fillna(0)

    def calc_units(start, end, c3, c6_set, must_inc, must_exc):
        s, e = pd.Timestamp(start), pd.Timestamp(f'{end} 23:59:59')
        sub = d[(d['單據日期'] >= s) & (d['單據日期'] <= e)]
        # 認證機（品牌 881/885/886/888）無論 C3 為何，只要符合條件就計入
        brand_sub = _brand.reindex(sub.index).fillna(0)
        c3_mask = (sub['類別3代碼'] == c3) | brand_sub.isin(_cert)
        if c6_set is None:
            # 名稱辨別模式：只用 C3 + must_inc/must_exc（不限 C6），適用 MacBook 各型號
            sub = sub[c3_mask]
        else:
            sub = sub[c3_mask & sub['類別6代碼'].isin(c6_set)]
        if must_inc: sub = sub[sub['名稱'].str.contains(must_inc, case=False, na=False)]
        if must_exc: sub = sub[~sub['名稱'].str.contains(must_exc, case=False, na=False)]
        # 台數邏輯：尾款當天計入（銷售完成），訂金不計，銷退扣除
        sale = sub.loc[sub['交易類型'].isin(SALE_TYPES), '數量'].sum()
        ret  = sub.loc[sub['交易類型'] == '銷退', '數量'].abs().sum()
        return int(sale - ret)

    prev_end   = quarter_start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=6)

    weeks = [(quarter_start + timedelta(weeks=i),
              quarter_start + timedelta(weeks=i, days=6)) for i in range(13)]

    table = {}
    for row_idx, (c3, c6_set, inc, exc) in PRODUCT_ROWS.items():
        prev = calc_units(prev_start, prev_end, c3, c6_set, inc, exc)
        weekly = []
        for ws_, we_ in weeks:
            if ws_ > week_end:
                weekly.append(0)
            else:
                weekly.append(calc_units(ws_, min(we_, week_end), c3, c6_set, inc, exc))
        table[row_idx] = [prev] + weekly

    for st_row, members in SUBTOTAL_ROWS.items():
        combined = [0] * 14
        for r in members:
            if r in table:
                for i, v in enumerate(table[r]):
                    combined[i] += v
        table[st_row] = combined

    # Excel row = table row + 1 (header is row 1)
    # Excel cols: B=2(prev W13), C=3(W01)...O=15(W13), P=16(Total)
    for row_idx, vals in table.items():
        excel_row = row_idx + 1
        total = sum(vals[1:])
        ws.cell(row=excel_row, column=2).value = vals[0] or None   # prev W13
        for wi, v in enumerate(vals[1:], start=3):
            ws.cell(row=excel_row, column=wi).value = v or None    # W01-W13
        ws.cell(row=excel_row, column=16).value = total or None    # Total

# ─── Sheet 2: 門市週報 ────────────────────────────────────────────────────────
def fill_sheet2(ws, df_cur, df_prev, sacare_prices, dates: dict, traffic=None, emp_count=None):
    print('  Sheet 2: 門市週報', flush=True)

    def get(df, start, end):
        return calc_metrics(period(df, start, end), sacare_prices)

    wk_prev = get(df_cur, dates['prev_wk_start'], dates['prev_wk_end'])
    wk      = get(df_cur, dates['wk_start'],      dates['wk_end'])
    mo      = get(df_cur, dates['mo_start'],       dates['mo_end'])
    lm      = get(df_cur, dates['lm_start'],       dates['lm_end'])
    ly      = get(df_prev, dates['ly_start'],      dates['ly_end'])

    # Column: B=2(上週), C=3(本週), F=6(本月), I=9(上月同期), L=12(去年同期)
    cols = {'上週': 2, '本週': 3, '本月': 6, '上月': 9, '去年': 12}
    data = {'上週': wk_prev, '本週': wk, '本月': mo, '上月': lm, '去年': ly}

    def sv(row, col_name, metric, pct=False):
        v = data[col_name].get(metric, 0)
        ws.cell(row=row, column=cols[col_name]).value = round(v, 4) if pct else int(v) if v else None

    def pct(row, col_name, num_key, den_key):
        m = data[col_name]
        num, den = m.get(num_key, 0), m.get(den_key, 0)
        ws.cell(row=row, column=cols[col_name]).value = safe_rate(num, den) if den else None

    for col in cols:
        m = data[col]
        sv(2,  col, 'total_rev')
        sv(3,  col, 'total_rev')           # 零售 = 總 (no 專案)
        ws.cell(row=4, column=cols[col]).value = None  # 專案 = 0
        sv(5,  col, 'rev_3001')
        sv(6,  col, 'rev_3002')
        sv(7,  col, 'rev_3003')
        pct(8, col, 'rev_3003', 'total_rev')           # 3PP搭售率
        sv(9,  col, 'sa_rev')
        pct(10, col, 'sa_rev', 'total_rev')             # SA搭售率
        sv(11, col, 'total_gross')
        sv(12, col, 'apl_gross')
        sv(13, col, 'tpp_gross')
        sv(14, col, 'sa_gross')
        sv(15, col, 'cpu_non_mini')
        sv(16, col, 'cpu_mini')
        sv(17, col, 'cpu_total')
        sv(18, col, 'acpp_mac')
        pct(19, col, 'acpp_mac', 'cpu_total')           # ACPP-MAC搭售率
        sv(20, col, 'sa_cpu')
        pct(21, col, 'sa_cpu', 'cpu_total')             # SA CPU搭售率
        sv(22, col, 'ipad')
        sv(23, col, 'sa_ipad')
        pct(24, col, 'sa_ipad', 'ipad')                 # SA iPad搭售率
        sv(25, col, 'iphone')
        sv(26, col, 'sa_iphone')
        pct(27, col, 'sa_iphone', 'iphone')             # SA iPhone搭售率
        sv(28, col, 'watch')
        sv(29, col, 'sa_watch')
        pct(30, col, 'sa_watch', 'watch')               # SA Watch搭售率
        # Row 31 人均產值 = 總營業額 / 總員工數（編制人數，前端輸入）
        if emp_count:
            ws.cell(row=31, column=cols[col]).value = int(m.get('total_rev', 0) / emp_count)
        # Row 32 來客數（ShopperTrak 人流；查不到則留空）
        visitors = (traffic or {}).get(col)
        if visitors is not None:
            ws.cell(row=32, column=cols[col]).value = int(visitors)
        sv(33, col, 'txn_count')
        # Row 34 提袋率 = 成交筆數 / 來客數（需有來客數）
        if visitors:
            ws.cell(row=34, column=cols[col]).value = safe_rate(m.get('txn_count', 0), visitors)

    # ── 計算差異欄位 D/E/H/J/K/M/N (全部用公式重算) ──────────────────────────
    # Active rows: the ones actually written by sv()/pct()
    active_rows = list(range(2, 31)) + [33]
    if emp_count:
        active_rows.append(31)                                   # 人均產值
    if traffic and any(v is not None for v in traffic.values()):
        active_rows += [32, 34]                                  # 來客數、提袋率
    for r in active_rows:
        B = ws.cell(row=r, column=2).value
        C = ws.cell(row=r, column=3).value
        F = ws.cell(row=r, column=6).value
        G = ws.cell(row=r, column=7).value   # 本月目標 (user-filled)
        I = ws.cell(row=r, column=9).value
        L = ws.cell(row=r, column=12).value
        # D = C - B (週差異)
        if C is not None and B is not None:
            ws.cell(row=r, column=4).value = round(C - B, 6)
        else:
            ws.cell(row=r, column=4).value = None
        # E = D/B (週差異%)
        ws.cell(row=r, column=5).value = safe_rate((C or 0) - (B or 0), B)
        # H = F/G (達成率) — G might be a string ('-') for 專案 row
        G_num = G if isinstance(G, (int, float)) else None
        ws.cell(row=r, column=8).value = safe_rate(F, G_num)
        # J = F - I (上月同期差異)
        if F is not None and I is not None:
            ws.cell(row=r, column=10).value = round(F - I, 6)
        else:
            ws.cell(row=r, column=10).value = None
        # K = J/I (上月同期差異%)
        ws.cell(row=r, column=11).value = safe_rate((F or 0) - (I or 0), I)
        # M = F - L (去年同期差異)
        if F is not None and L is not None:
            ws.cell(row=r, column=13).value = round(F - L, 6)
        else:
            ws.cell(row=r, column=13).value = None
        # N = M/L (去年同期差異%)
        ws.cell(row=r, column=14).value = safe_rate((F or 0) - (L or 0), L)

# ─── Sheet 3: 3PP配件比較 ──────────────────────────────────────────────────────
def fill_sheet3(ws, df_cur, df_prev, sacare_prices, dates: dict):
    print('  Sheet 5: 3PP配件比較', flush=True)
    sa_codes = set(sacare_prices.keys())

    def c4_rev(df, start, end, c4_code):
        d = period(df, start, end)
        d = d[~d['存貨代碼'].astype(str).str.strip().isin(sa_codes)]
        d = d[(d['類別3代碼'] == 3003.0) & (d['類別4代碼'] == c4_code)]
        return int(d['NET'].sum())

    def sa_rev_total(df, start, end):
        d = period(df, start, end)
        sa = d[d['存貨代碼'].astype(str).str.strip().isin(sa_codes)].copy()
        sa['SA_NET'] = sa['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa['數量'].fillna(0)
        return int(sa['SA_NET'].sum())

    # Columns: C=3(上週), D=4(本週), G=7(本月), H=8(上月同期), K=11(去年同期)
    #          N=14(2025 YTD), O=15(2026 YTD)
    periods = {
        'prev_wk': (df_cur,  dates['prev_wk_start'], dates['prev_wk_end']),
        'wk':      (df_cur,  dates['wk_start'],       dates['wk_end']),
        'mo':      (df_cur,  dates['mo_start'],        dates['mo_end']),
        'lm':      (df_cur,  dates['lm_start'],        dates['lm_end']),
        'ly':      (df_prev, dates['ly_start'],        dates['ly_end']),
        'ytd_cur': (df_cur,  dates['ytd_cur_start'],   dates['mo_end']),
        'ytd_prv': (df_prev, dates['ytd_prv_start'],   dates['ly_end']),
    }

    col_map = {'prev_wk': 3, 'wk': 4, 'mo': 7, 'lm': 8,
               'ly': 11, 'ytd_prv': 14, 'ytd_cur': 15}

    def write_diff_cols(ws, row):
        """Compute and write difference/percentage columns E/F/I/J/L/M/P/Q."""
        C = ws.cell(row=row, column=3).value or 0
        D = ws.cell(row=row, column=4).value or 0
        G = ws.cell(row=row, column=7).value or 0
        H = ws.cell(row=row, column=8).value or 0
        K = ws.cell(row=row, column=11).value or 0
        N = ws.cell(row=row, column=14).value or 0
        O = ws.cell(row=row, column=15).value or 0
        # E/F: 本週 vs 上週
        ws.cell(row=row, column=5).value  = (D - C) or None
        ws.cell(row=row, column=6).value  = safe_rate(D - C, C)
        # I/J: 本月 vs 上月同期
        ws.cell(row=row, column=9).value  = (G - H) or None
        ws.cell(row=row, column=10).value = safe_rate(G - H, H)
        # L/M: 本月 vs 去年同期
        ws.cell(row=row, column=12).value = (G - K) or None
        ws.cell(row=row, column=13).value = safe_rate(G - K, K)
        # P/Q: 今年YTD vs 去年YTD
        ws.cell(row=row, column=16).value = (O - N) or None
        ws.cell(row=row, column=17).value = safe_rate(O - N, N)

    for c4, excel_row in C4_ROWS.items():
        for pname, (df, s, e) in periods.items():
            v = c4_rev(df, s, e, c4)
            ws.cell(row=excel_row, column=col_map[pname]).value = v or None
        write_diff_cols(ws, excel_row)

    # SA CARE row (row 16)
    sa_row = 16
    for pname, (df, s, e) in periods.items():
        v = sa_rev_total(df, s, e)
        ws.cell(row=sa_row, column=col_map[pname]).value = v or None
    write_diff_cols(ws, sa_row)

    # Row 17 (加總): sum rows 2-16 (all C4 + SA Care)
    # Row 18 (純配件): sum rows 2-15 (C4 only, exclude SA Care row 16)
    for value_col in [3, 4, 7, 8, 11, 14, 15]:
        total_all  = sum(ws.cell(row=r, column=value_col).value or 0 for r in range(2, 17))
        total_pure = sum(ws.cell(row=r, column=value_col).value or 0 for r in range(2, 16))
        ws.cell(row=17, column=value_col).value = total_all  or None
        ws.cell(row=18, column=value_col).value = total_pure or None
    write_diff_cols(ws, 17)
    write_diff_cols(ws, 18)

# ─── Sheet 4/5: 銷售排名 ──────────────────────────────────────────────────────
def fill_sheet45(ws4, ws5, df_cur, sacare_prices, dates: dict):
    print('  Sheet 6/7: 銷售排名', flush=True)
    sa_codes = set(sacare_prices.keys())

    def get_ranking(df, start, end, vap_only=False):
        d = period(df, start, end)
        d = d[~d['存貨代碼'].astype(str).str.strip().isin(sa_codes)]
        d = d[d['類別3代碼'] == 3003.0]
        # Exclude gift (折扣 >= 80 or 金額 == 0)
        is_return = d['交易類型'] == '銷退'
        is_gift = (d.get('折扣', pd.Series(0, index=d.index)).fillna(0) >= 80) | \
                  ((d['NET'].abs() == 0) & ~is_return)
        d = d[~is_gift]
        if vap_only:
            d = d[d['品牌代碼'].isin(VAP_BRANDS)]
        else:
            d = d[~d['品牌代碼'].isin(VAP_BRANDS)]

        grouped = d.groupby('存貨代碼').agg(
            品名=('名稱', 'first'),
            qty=('數量', 'sum'),
            rev=('NET', 'sum')
        ).reset_index()
        return grouped.sort_values('qty', ascending=False).head(10)

    def write_ranking(ws, wk_rank, mo_rank):
        for i in range(10):
            row = i + 2
            # 週銷 (cols B=2, C=3, D=4)
            if i < len(wk_rank):
                r = wk_rank.iloc[i]
                ws.cell(row=row, column=2).value = str(r['存貨代碼'])
                ws.cell(row=row, column=3).value = str(r['品名'])
                ws.cell(row=row, column=4).value = int(r['qty'])
            else:
                for c in [2, 3, 4]:
                    ws.cell(row=row, column=c).value = None
            # 月銷 (cols E=5, F=6, G=7, H=8)
            if i < len(mo_rank):
                r = mo_rank.iloc[i]
                ws.cell(row=row, column=5).value = str(r['存貨代碼'])
                ws.cell(row=row, column=6).value = str(r['品名'])
                ws.cell(row=row, column=7).value = int(r['qty'])
                ws.cell(row=row, column=8).value = int(r['rev'])
            else:
                for c in [5, 6, 7, 8]:
                    ws.cell(row=row, column=c).value = None

    wk_3pp = get_ranking(df_cur, dates['wk_start'], dates['wk_end'], vap_only=False)
    mo_3pp = get_ranking(df_cur, dates['mo_start'], dates['mo_end'], vap_only=False)
    write_ranking(ws4, wk_3pp, mo_3pp)

    wk_vap = get_ranking(df_cur, dates['wk_start'], dates['wk_end'], vap_only=True)
    mo_vap = get_ranking(df_cur, dates['mo_start'], dates['mo_end'], vap_only=True)
    write_ranking(ws5, wk_vap, mo_vap)

# ─── Per-employee calculations ─────────────────────────────────────────────────
def calc_employee(df: pd.DataFrame, emp_code: str, sacare_prices: dict) -> dict:
    """Calculate all per-employee metrics for a filtered period."""
    d = df[df['員工代碼'] == emp_code].copy()
    sa_codes = set(sacare_prices.keys())
    non_sa = d[~d['存貨代碼'].astype(str).str.strip().isin(sa_codes)]

    # Revenue (excl SA)
    rev_excl_sa = int(non_sa['NET'].sum())
    rev_3003    = int(non_sa.loc[non_sa['類別3代碼'] == 3003.0, 'NET'].sum())
    rev_3002    = int(non_sa.loc[non_sa['類別3代碼'] == 3002.0, 'NET'].sum())  # Apple 原廠配件
    # D 欄：原廠商品營業額（不含 SAcare，尾款才紀錄，訂金不計）
    # 排除：C1∈{1002,1004,1008} / C3∈{3047,3003,3004,3018,3019,3012}
    #        C6∈{6888,6889} / 指定存貨代碼
    non_sa_d = non_sa[non_sa['交易類型'].isin(SALE_TYPES | {'銷退'})]
    _c1  = non_sa_d.get('類別1代碼', pd.Series(dtype=float, index=non_sa_d.index)).fillna(0)
    _c6  = non_sa_d.get('類別6代碼', pd.Series(dtype=float, index=non_sa_d.index)).fillna(0)
    _sku = non_sa_d['存貨代碼'].astype(str).str.strip()
    apple_mask = (
        ~_c1.isin(C1_EXCLUDED_FROM_APPLE) &
        ~non_sa_d['類別3代碼'].isin(C3_EXCLUDED_FROM_APPLE) &
        ~_c6.isin(C6_EXCLUDED_FROM_APPLE) &
        ~_sku.isin(SKU_EXCLUDED_FROM_APPLE)
    )
    rev_apple = int(non_sa_d.loc[apple_mask, 'NET'].sum())

    # SAcare：只計 SALE_TYPES（銷售/尾款），排除訂金避免雙重計算
    sa_rows = d[d['存貨代碼'].astype(str).str.strip().isin(sa_codes)].copy()
    sa_sold = sa_rows[sa_rows['交易類型'].isin(SALE_TYPES)]
    sa_ret  = sa_rows[sa_rows['交易類型'] == '銷退']
    sa_sold_net = (sa_sold['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa_sold['數量'].fillna(0)).sum()
    sa_ret_net  = (sa_ret['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa_ret['數量'].abs()).sum()
    sa_rev = int(sa_sold_net - sa_ret_net)

    # Gross profit (Sheet 6 H/I 未稅毛利)
    # H 原廠毛利：依 ERP「13-門市獎金Apple毛利額未稅-員工」
    #   C3 白名單 {3001,3002,3032,3033,3046} + 99901689（抵用券兌換, C3=NaN）
    #   交易類型：包含 銷售/訂金/銷退，排除 尾款/退訂
    #   原因：800AB 訂金列記錄原廠商品全額未稅收入；尾款列收入=0
    # I 3PP毛利：依 ERP「14-門市獎金3PP毛利額未稅-員工」
    #   C3 白名單 {3003,3006,3004,3018,3019,3012}
    #   交易類型：排除 尾款（訂金/退訂/銷退 皆納入）
    #   原因：ERP 以訂金發生時點認列 3PP 收入與成本；尾款列為 3PP 贈品成本補登列
    #         (NET=0，cost>0)，ERP 未計入；退訂會自動與原訂金相抵。
    #   v1.1.15 修正：改為只排除 尾款（原先錯誤地排除 訂金/退訂）
    # 公式：淨銷售金額(未稅) - 單位成本 × 數量（不加銷退金額(未稅)，不乘/除 1.05）
    SA_CARE_GROSS_EXCL = {
        '99903303','99903302','99200168','99500006','99900946','99900947','99900948',
        '99900949','99900950','99901684','99901685','99902607','99902608','99902609',
        '99902610','99903343','99903339',
        '99200202',  # 教育價活動（內部代碼，非實際商品，已排除於 D 欄）
        '99200201',  # 促銷組合（內部代碼）
        '99500203',  # v1.1.16: 預收訂金（會計科目，非實際商品，未來由商品銷售/尾款列認列）
    }
    # Apple-specific additional exclusions (800AB format, leading zeros removed)
    SA_APPLE_GROSS_EXCL = SA_CARE_GROSS_EXCL | {
        '7307154','7309136','7309137','7310037','7310042',
        '7310053','7310093','7311242',
        '88600895','88601027','90501795','90501799',
    }

    sku_str = d['存貨代碼'].astype(str).str.strip()
    c6_excl = d.get('類別6代碼', pd.Series(dtype=float, index=d.index)).isin([6888.0, 6889.0])
    c1_excl = d.get('類別1代碼', pd.Series(dtype=float, index=d.index)).isin([1002.0, 1004.0, 1008.0])
    excl_apl = sku_str.isin(SA_APPLE_GROSS_EXCL) | c6_excl | c1_excl
    excl_tpp = sku_str.isin(SA_CARE_GROSS_EXCL)  | c6_excl | c1_excl

    apl_mask = d['類別3代碼'].isin([3001.0, 3002.0, 3032.0, 3033.0, 3046.0]) | (sku_str == '99901689')

    def gross_apl():
        sub = d[apl_mask & ~excl_apl & ~d['交易類型'].isin(['尾款', '退訂'])]
        net_ux = sub.get('淨銷售金額(未稅)', pd.Series(0, index=sub.index)).fillna(0)
        # C3=3032 (ACPP 代收保費): 800AB 將含稅金額存入 淨銷售金額(未稅) 欄位，
        # 而 ERP 銷貨收入(未稅) 使用含稅÷1.05 的真正未稅金額，需修正。
        is_acpp = sub['類別3代碼'] == 3032.0
        net_ux = net_ux.where(~is_acpp, net_ux / 1.05)
        cost   = sub.get('單位成本', pd.Series(0, index=sub.index)).fillna(0) * sub['數量'].fillna(0)
        return int((net_ux - cost).sum())

    tpp_mask = d['類別3代碼'].isin([3003.0, 3006.0, 3004.0, 3018.0, 3019.0, 3012.0])

    def gross_tpp():
        sub = d[tpp_mask & ~excl_tpp & ~d['交易類型'].isin(['尾款'])]
        net_ux = sub.get('淨銷售金額(未稅)', pd.Series(0, index=sub.index)).fillna(0)
        cost   = sub.get('單位成本', pd.Series(0, index=sub.index)).fillna(0) * sub['數量'].fillna(0)
        return int((net_ux - cost).sum())

    apl_gross = gross_apl()
    tpp_gross = gross_tpp()

    # 認證機品牌（無論 C3 為何，只要 C6/C4 符合就計入）
    _cert  = {881.0, 885.0, 886.0, 888.0}
    _brand = d.get('品牌代碼', pd.Series(dtype=float, index=d.index)).fillna(0)

    # Unit counts
    # Mac (C6): need to cover all Mac C6 codes across model years
    def net_units_c6(c6_set, c3=None):
        m = d['類別6代碼'].isin(c6_set)
        if c3:
            m &= (d['類別3代碼'] == c3) | _brand.isin(_cert)
        sale = d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum()
        ret  = d.loc[m & (d['交易類型'] == '銷退'),  '數量'].abs().sum()
        return int(sale - ret)

    # iPhone / Watch / iPad (C4): stable across model years + 認證機品牌 bypass
    def net_units_c4(c4_set, c3=3001.0):
        m = ((d['類別3代碼'] == c3) | _brand.isin(_cert)) & d['類別4代碼'].isin(c4_set)
        sale = d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum()
        ret  = d.loc[m & (d['交易類型'] == '銷退'),  '數量'].abs().sum()
        return int(sale - ret)

    cpu     = net_units_c6(C6_CPU, c3=3001.0)   # 引用 C6_CPU 常數，新世代只需更新 C6_CPU
    iphone  = net_units_c4(C4_IPHONE)
    ipad    = net_units_c4(C4_IPAD)
    watch   = net_units_c4(C4_WATCH)
    airpods = net_units_c6(C6_AIRPODS, c3=3002.0)

    # ACPP+ = 類別3=3032 (AppleCare+ 代收保費), grouped by device type via product name
    ac_rows = d[(d['類別3代碼'] == 3032.0) & d['交易類型'].isin(SALE_TYPES)].copy()
    ac_ret  = d[(d['類別3代碼'] == 3032.0) & (d['交易類型'] == '銷退')].copy()
    def acpp_name(keyword):
        kw = keyword.lower()
        sold = ac_rows[ac_rows['名稱'].astype(str).str.lower().str.contains(kw, na=False)]['數量'].sum()
        retn = ac_ret[ ac_ret['名稱'].astype(str).str.lower().str.contains(kw, na=False)]['數量'].abs().sum()
        return int(sold - retn)

    mac_acpp     = acpp_name('mac')
    ipad_acpp    = acpp_name('ipad')
    iphone_acpp  = acpp_name('iphone')
    watch_acpp   = acpp_name('watch')
    airpods_acpp = acpp_name('airpods')

    # SAcare counts by device
    def sa_cnt(c6_set):
        m = d['類別6代碼'].isin(c6_set) & d['存貨代碼'].astype(str).str.strip().isin(sa_codes)
        return int(d.loc[m & d['交易類型'].isin(SALE_TYPES), '數量'].sum() -
                   d.loc[m & (d['交易類型'] == '銷退'),  '數量'].abs().sum())

    mac_sa    = sa_cnt(C6_SA['cpu'])
    ipad_sa   = sa_cnt(C6_SA['ipad'])
    iphone_sa = sa_cnt(C6_SA['iphone'])
    watch_sa  = sa_cnt(C6_SA['watch'])
    airpods_sa = sa_cnt(C6_SA['airpods'])

    # C4 3PP accessory revenues
    def acc_rev(c4_set):
        m = d['類別4代碼'].isin(c4_set) & (d['類別3代碼'] == 3003.0)
        return int(d.loc[m, 'NET'].sum())

    vap_rev    = int(d.loc[d['品牌代碼'].isin(VAP_BRANDS), 'NET'].sum())
    _office_mask = d['名稱'].str.lower().str.contains('office|microsoft', na=False)
    office_qty = int(d.loc[_office_mask & d['交易類型'].isin(SALE_TYPES), '數量'].sum()) - \
                 int(d.loc[_office_mask & (d['交易類型'] == '銷退'), '數量'].abs().sum())

    return dict(
        rev_excl_sa=rev_excl_sa, rev_3003=rev_3003, rev_3002=rev_3002, rev_apple=rev_apple,
        sa_rev=sa_rev, apl_gross=apl_gross, tpp_gross=tpp_gross,
        cpu=cpu, iphone=iphone, ipad=ipad, watch=watch, airpods=airpods,
        mac_acpp=mac_acpp, ipad_acpp=ipad_acpp,
        iphone_acpp=iphone_acpp, watch_acpp=watch_acpp, airpods_acpp=airpods_acpp,
        mac_sa=mac_sa, ipad_sa=ipad_sa, iphone_sa=iphone_sa,
        watch_sa=watch_sa, airpods_sa=airpods_sa,
        cpu_acc=acc_rev(C4_CPU_SET), iphone_acc=acc_rev(C4_IPHONE_SET),
        ipad_acc=acc_rev(C4_IPAD_SET), watch_acc=acc_rev(C4_WATCH_SET),
        airpods_acc=acc_rev(C4_AIRPODS_SET),
        cpu_ios_acc=acc_rev({4012.0}),   # CPU/iOS通用週邊
        ios_acc=acc_rev({4022.0}),        # iOS通用週邊配件
        vap_rev=vap_rev, office_qty=office_qty,
    )

# ─── Sheet 6: 個人新制獎金 ────────────────────────────────────────────────────
def fill_sheet6(ws, df_cur, sacare_prices, dates: dict):
    print('  Sheet 10: 個人新制獎金', flush=True)
    d_mo = period(df_cur, dates['mo_start'], dates['mo_end'])
    # 排除等級代碼 05（非全職／兼職員工，不納入獎金計算）
    if '等級代碼' in d_mo.columns:
        d_mo = d_mo[d_mo['等級代碼'] != '05']

    for i, (code, _) in enumerate(EMPLOYEES):
        row = i + 2
        m = calc_employee(d_mo, code, sacare_prices)
        ws.cell(row=row, column=2).value  = m['rev_excl_sa'] or None   # B 營業額
        ws.cell(row=row, column=3).value  = m['rev_3003'] or None       # C 3PP 營業額
        ws.cell(row=row, column=4).value  = m['rev_apple'] or None       # D 原廠商品營業額（排除3003/3004/3012/3018/3019/3047）
        ws.cell(row=row, column=5).value  = m['sa_rev'] or None         # E SA Care
        # 未稅毛利由 calc_employee 直接依 ERP 公式計算（不需 /1.05）
        apl_gross_ex = m['apl_gross'] or None
        tpp_gross_ex = m['tpp_gross'] or None
        ws.cell(row=row, column=8).value  = apl_gross_ex                # H 原廠毛利(未稅)
        ws.cell(row=row, column=9).value  = tpp_gross_ex                # I 3PP 毛利(未稅)
        # F, G, J, K = Excel formulas — skip

    # 加總列：動態掃描 A 欄找「加總」文字，避免因模板結構差異（如有無總公司列）而寫錯列
    emp_end    = 1 + len(EMPLOYEES)          # 最後一位員工的列號
    emp_rows_6 = range(2, emp_end + 1)
    # 從員工結束列之後向下找「加總」
    total_row = None
    for r in range(emp_end + 1, emp_end + 10):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and '加總' in str(cell_val):
            total_row = r
            break
    if total_row is None:
        total_row = emp_end + 1              # fallback：緊接在最後員工後面
    for col in [2, 3, 4, 5, 8, 9]:
        total = sum(ws.cell(row=r, column=col).value or 0 for r in emp_rows_6)
        ws.cell(row=total_row, column=col).value = total or None

# ─── Sheet 7/8: 個人週/月主機 ─────────────────────────────────────────────────
def fill_sheet78(ws7, ws8, df_cur, sacare_prices, dates: dict):
    print('  Sheet 11/12: 個人週/月主機', flush=True)
    d_wk = period(df_cur, dates['wk_start'], dates['wk_end'])
    d_mo = period(df_cur, dates['mo_start'], dates['mo_end'])

    def write_emp_row(ws, row, m):
        cpu, ipad, iphone, watch, airpods = m['cpu'], m['ipad'], m['iphone'], m['watch'], m['airpods']
        mac_acpp, mac_sa = m['mac_acpp'], m['mac_sa']
        ipad_acpp, ipad_sa = m['ipad_acpp'], m['ipad_sa']
        iphone_acpp, iphone_sa = m['iphone_acpp'], m['iphone_sa']
        watch_acpp, watch_sa = m['watch_acpp'], m['watch_sa']
        airpods_acpp, airpods_sa = m['airpods_acpp'], m['airpods_sa']

        mac_rate    = safe_rate(mac_acpp + mac_sa, cpu)
        ipad_rate   = safe_rate(ipad_acpp + ipad_sa, ipad)
        iphone_rate = safe_rate(iphone_acpp + iphone_sa, iphone)
        ios_rate    = safe_rate(iphone_acpp + iphone_sa + ipad_acpp + ipad_sa, iphone + ipad)
        watch_rate  = safe_rate(watch_acpp + watch_sa, watch)
        ap_rate     = safe_rate(airpods_acpp + airpods_sa, airpods)

        vals = [
            cpu or None, mac_acpp or None, mac_sa or None, mac_rate,
            ipad or None, ipad_acpp or None, ipad_sa or None, ipad_rate,
            iphone or None, iphone_acpp or None, iphone_sa or None, iphone_rate,
            ios_rate,
            watch or None, watch_acpp or None, watch_sa or None, watch_rate,
            airpods or None, airpods_acpp or None, airpods_sa or None, ap_rate,
        ]
        for col_offset, v in enumerate(vals):
            ws.cell(row=row, column=col_offset + 2).value = v

    for i, (code, _) in enumerate(EMPLOYEES):
        row = i + 3  # data starts at row 3
        write_emp_row(ws7, row, calc_employee(d_wk, code, sacare_prices))
        write_emp_row(ws8, row, calc_employee(d_mo, code, sacare_prices))

    # ── 加總列動態：起始 row 3，加總列 = 3 + 員工人數 ────────────────────────
    def write_total_row(ws):
        total_row = 3 + len(EMPLOYEES)
        emp_rows  = range(3, total_row)
        # Count columns: B(2),C(3),D(4), F(6),G(7),H(8), J(10),K(11),L(12), O(15),P(16),Q(17), S(19),T(20),U(21)
        count_cols = [2,3,4, 6,7,8, 10,11,12, 15,16,17, 19,20,21]
        totals = {}
        for col in count_cols:
            totals[col] = sum(ws.cell(row=r, column=col).value or 0 for r in emp_rows)
            ws.cell(row=total_row, column=col).value = totals[col] or None
        # Rate columns recomputed from totals
        # E(5): Mac rate = (C+D)/B
        ws.cell(row=total_row, column=5).value  = safe_rate(totals[3]+totals[4], totals[2])
        # I(9): iPad rate = (G+H)/F
        ws.cell(row=total_row, column=9).value  = safe_rate(totals[7]+totals[8], totals[6])
        # M(13): iPhone rate = (K+L)/J
        ws.cell(row=total_row, column=13).value = safe_rate(totals[11]+totals[12], totals[10])
        # N(14): iOS rate = (K+L+G+H)/(J+F)
        ws.cell(row=total_row, column=14).value = safe_rate(totals[11]+totals[12]+totals[7]+totals[8], totals[10]+totals[6])
        # R(18): Watch rate = (P+Q)/O
        ws.cell(row=total_row, column=18).value = safe_rate(totals[16]+totals[17], totals[15])
        # V(22): AirPods rate = (T+U)/S
        ws.cell(row=total_row, column=22).value = safe_rate(totals[20]+totals[21], totals[19])

    write_total_row(ws7)
    write_total_row(ws8)

# ─── Sheet 9: 個人月3PP ───────────────────────────────────────────────────────
def fill_sheet9(ws, df_cur, sacare_prices, dates: dict):
    print('  Sheet 13: 個人月3PP', flush=True)
    d_mo = period(df_cur, dates['mo_start'], dates['mo_end'])

    for i, (code, _) in enumerate(EMPLOYEES):
        row = i + 2
        m = calc_employee(d_mo, code, sacare_prices)
        ws.cell(row=row, column=2).value  = m['cpu_acc']     or None  # B CPU配件
        ws.cell(row=row, column=3).value  = m['iphone_acc']  or None  # C iPhone配件
        ws.cell(row=row, column=4).value  = m['ipad_acc']    or None  # D iPad配件
        ws.cell(row=row, column=5).value  = m['cpu_ios_acc'] or None  # E CPU/iOS通用週邊
        ws.cell(row=row, column=6).value  = m['ios_acc']     or None  # F iOS通用週邊配件
        ws.cell(row=row, column=7).value  = m['watch_acc']   or None  # G Watch配件
        ws.cell(row=row, column=8).value  = m['airpods_acc'] or None  # H AirPods配件
        ws.cell(row=row, column=9).value  = m['vap_rev']     or None  # I VAP
        ws.cell(row=row, column=10).value = m['office_qty']  or None  # J Office數量

    # 加總列動態：起始 row 2，加總列 = 2 + 員工人數
    total_row  = 2 + len(EMPLOYEES)
    emp_rows_9 = range(2, total_row)
    for col in range(2, 11):   # B～J
        total = sum(ws.cell(row=r, column=col).value or 0 for r in emp_rows_9)
        ws.cell(row=total_row, column=col).value = total or None

# ─── Sheet 3/4: 每月重點 / Speakers ───────────────────────────────────────────
def _net_qty(d: pd.DataFrame, mask) -> int:
    """淨數量：銷售/尾款 正計、銷退 取絕對值扣回（同 calc_employee 的 net_units_* 慣例）。"""
    sale = d.loc[mask & d['交易類型'].isin(SALE_TYPES), '數量'].sum()
    ret  = d.loc[mask & (d['交易類型'] == '銷退'), '數量'].abs().sum()
    return int(sale - ret)

def _focus_mask(d: pd.DataFrame):
    return d['品牌代碼'].isin(MONTHLY_FOCUS_BRANDS)

def _speakers_mask(d: pd.DataFrame):
    return (d['類別3代碼'] == 3003.0) & d['類別4代碼'].isin(SPEAKERS_C4)

def _label_row(ws, label: str, start: int = 2):
    """在 A 欄由 start 列往下找標籤所在列（員工列數為動態，不能寫死列號）。"""
    for r in range(start, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == label:
            return r
    return None

def fill_sheet_focus(ws, df_cur, dates: dict):
    """第 3 頁 每月重點：C=週銷售數、D=月銷售數（品牌 MONTHLY_FOCUS_BRANDS 的淨數量）。"""
    print('  Sheet 3: 每月重點', flush=True)
    d_wk = period(df_cur, dates['wk_start'], dates['wk_end'])
    d_mo = period(df_cur, dates['mo_start'], dates['mo_end'])
    total_row = _label_row(ws, 'Total') or (2 + len(EMPLOYEES))

    for i, (code, _) in enumerate(EMPLOYEES):
        row = i + 2
        for col, d in ((3, d_wk), (4, d_mo)):
            e = d[d['員工代碼'] == code]
            ws.cell(row=row, column=col).value = _net_qty(e, _focus_mask(e)) or None

    for col in (3, 4):
        total = sum(ws.cell(row=r, column=col).value or 0 for r in range(2, total_row))
        ws.cell(row=total_row, column=col).value = total or None

def fill_sheet_speakers(ws, df_cur, dates: dict):
    """第 4 頁 Speakers（3PP 藍牙喇叭 C3=3003 且 C4∈SPEAKERS_C4）：
    C/D=本週銷售數/金額、E/F=月銷售數/金額，另加一格全店年累積金額。"""
    print('  Sheet 4: Speakers', flush=True)
    d_wk = period(df_cur, dates['wk_start'], dates['wk_end'])
    d_mo = period(df_cur, dates['mo_start'], dates['mo_end'])
    total_row = _label_row(ws, 'Total') or (2 + len(EMPLOYEES))

    for i, (code, _) in enumerate(EMPLOYEES):
        row = i + 2
        for (qty_col, amt_col), d in (((3, 4), d_wk), ((5, 6), d_mo)):
            e = d[d['員工代碼'] == code]
            m = _speakers_mask(e)
            ws.cell(row=row, column=qty_col).value = _net_qty(e, m) or None
            ws.cell(row=row, column=amt_col).value = int(e.loc[m, 'NET'].sum()) or None

    for col in range(3, 7):
        total = sum(ws.cell(row=r, column=col).value or 0 for r in range(2, total_row))
        ws.cell(row=total_row, column=col).value = total or None

    # 年累積金額（全店，不限報表列出的員工）：今年 1/1 ～ 截止日。
    # 截止日＝前端「年對年截止日」有填就用它（與第 14/15 頁同步），留空則用本月結束日。
    ytd_end = dates.get('ytd_end') or dates['mo_end']
    d_ytd = period(df_cur, date(ytd_end.year, 1, 1), ytd_end)
    ytd_row = _label_row(ws, '年累積金額', start=total_row + 1)
    if ytd_row:
        amt = int(d_ytd.loc[_speakers_mask(d_ytd), 'NET'].sum())
        ws.cell(row=ytd_row, column=2).value = amt or None

# ─── Sheet 12/13 共用：YOY 累積區間 ───────────────────────────────────────────
def _yoy_periods(dates: dict) -> 'tuple[date, date, date, date]':
    """回傳 (今年起始, 今年截止, 去年起始, 去年截止)。
    截止日預設為本週末日；若 dates 內有 'yoy_end'（前端自訂年對年截止日）則優先採用。
    今年 = 截止日當年 1/1 ～ 截止日；去年 = 前一年 1/1 ～ 去年同月同日
    （2/29 等不存在時取當月最後一天）。"""
    end = dates.get('yoy_end') or dates['wk_end']
    cur_year = end.year
    prev_year = cur_year - 1
    last = monthrange(prev_year, end.month)[1]
    prv_e = date(prev_year, end.month, min(end.day, last))
    return date(cur_year, 1, 1), end, date(prev_year, 1, 1), prv_e


# ─── Sheet 10: 月報YOY（年對年累積比較）────────────────────────────────────────
def fill_sheet10(ws, df_cur, df_prev, sacare_prices, dates: dict, traffic=None, emp_count=None):
    print('  Sheet 14: 月報YOY', flush=True)
    # 累積區間：今年 1/1～截止日、去年 1/1～去年同日（截止日預設週末，可由前端自訂）
    cur_s, cur_e, prv_s, prv_e = _yoy_periods(dates)
    cur = calc_metrics(period(df_cur,  cur_s, cur_e), sacare_prices)
    prv = calc_metrics(period(df_prev, prv_s, prv_e), sacare_prices)

    # row -> (key, type)；type: m=金額/台數, p=比率(num,den), s=特例
    rows = {
        2:  ('total_rev', 'm'),                      # 總營業額
        3:  ('total_rev', 'm'),                      # 零售營業額（= 總，無專案）
        # row 4 專案營業額 → 留空（0）
        5:  ('rev_3001', 'm'),                        # Apple 主機營業額
        6:  ('rev_3002', 'm'),                        # Apple 配件營業額
        7:  ('rev_3003', 'm'),                        # 3PP配件營業額
        8:  (('rev_3003', 'total_rev'), 'p'),         # 3PP搭售率
        9:  ('sa_rev', 'm'),                          # SA Care 營業額
        10: (('sa_rev', 'total_rev'), 'p'),           # SA Care 搭售率
        11: ('total_gross', 'm'),                     # 總毛利額
        12: ('apl_gross', 'm'),                        # Apple 毛利額
        13: ('tpp_gross', 'm'),                        # 3PP 毛利額
        14: ('sa_gross', 'm'),                         # SA Care 毛利額
        15: ('cpu_ex_mini', 's'),                      # CPU 台數（不含Mac mini）= 總 − mini
        16: ('cpu_mini', 'm'),                         # Mac mini 台數
        17: ('cpu_total', 'm'),                        # CPU 總台數
        18: ('acpp_mac', 'm'),                         # ACPP-MAC 套數
        19: (('acpp_mac', 'cpu_total'), 'p'),          # ACPP-MAC 搭售率
        20: ('sa_cpu', 'm'),                           # SA Care for CPU 套數
        21: (('sa_cpu', 'cpu_total'), 'p'),            # SA Care for CPU 搭售率
        22: ('ipad', 'm'),                             # iPad 台數
        23: ('sa_ipad', 'm'),                          # SA Care for iPad 套數
        24: (('sa_ipad', 'ipad'), 'p'),                # SA Care for iPad 搭售率
        25: ('iphone', 'm'),                           # iPhone 台數
        26: ('sa_iphone', 'm'),                        # SA Care for iPhone 套數
        27: (('sa_iphone', 'iphone'), 'p'),            # SA Care for iPhone 搭售率
        28: ('watch', 'm'),                            # Watch 台數
        29: ('sa_watch', 'm'),                         # SA Care for Watch 套數
        30: (('sa_watch', 'watch'), 'p'),              # SA Care for Watch 搭售率
        # row 31 人均產值、32 來客數 → 來自 POS，非 800AB，留空
        33: ('txn_count', 'm'),                        # 成交筆數
    }

    def value(m, key, typ):
        if typ == 'p':
            n, d = key
            return safe_rate(m.get(n, 0), m.get(d, 0)) or 0
        if typ == 's' and key == 'cpu_ex_mini':
            return m['cpu_total'] - m['cpu_mini']   # MacBook + iMac
        return m.get(key, 0)

    for r, (key, typ) in rows.items():
        b = value(prv, key, typ)
        c = value(cur, key, typ)
        is_pct = (typ == 'p')
        ws.cell(row=r, column=2).value = (round(b, 4) if is_pct else int(b)) if (is_pct or b) else None  # B 去年同期累積
        ws.cell(row=r, column=3).value = (round(c, 4) if is_pct else int(c)) if (is_pct or c) else None  # C 今年累積
        # D 年差異 = 今年 − 去年
        diff = c - b
        ws.cell(row=r, column=4).value = round(diff, 6) if is_pct else (int(diff) if diff else None)
        # E 差異比例 = 年差異 / 去年
        ws.cell(row=r, column=5).value = safe_rate(diff, b)

    # row 31 人均產值（累積 = 累積總營業額 / 編制人數）、row 32 來客數（ShopperTrak）
    def write_yoy(row, b, c):
        ws.cell(row=row, column=2).value = int(b) if b else None
        ws.cell(row=row, column=3).value = int(c) if c else None
        d = (c or 0) - (b or 0)
        ws.cell(row=row, column=4).value = int(d) if d else None
        ws.cell(row=row, column=5).value = safe_rate(d, b)

    if emp_count:
        write_yoy(31, prv.get('total_rev', 0) / emp_count, cur.get('total_rev', 0) / emp_count)
    t = traffic or {}
    if t.get('prv') is not None or t.get('cur') is not None:
        write_yoy(32, t.get('prv') or 0, t.get('cur') or 0)


# ─── Sheet 11: 3PP YOY（各 3PP 類別年對年累積比較）─────────────────────────────
def fill_sheet11(ws, df_cur, df_prev, sacare_prices, dates: dict):
    print('  Sheet 15: 3PP YOY', flush=True)
    sa_codes = set(sacare_prices.keys())

    def c4_rev(df, start, end, c4_code):
        d = period(df, start, end)
        d = d[~d['存貨代碼'].astype(str).str.strip().isin(sa_codes)]
        d = d[(d['類別3代碼'] == 3003.0) & (d['類別4代碼'] == c4_code)]
        return int(d['NET'].sum())

    def sa_rev_total(df, start, end):
        d = period(df, start, end)
        sa = d[d['存貨代碼'].astype(str).str.strip().isin(sa_codes)].copy()
        sa['SA_NET'] = sa['存貨代碼'].astype(str).str.strip().map(sacare_prices) * sa['數量'].fillna(0)
        return int(sa['SA_NET'].sum())

    cur_s, cur_e, prv_s, prv_e = _yoy_periods(dates)

    def write_row(row, b, c):
        # C 去年度累積、D 今年度累積、E 差異金額、F 差異比例
        ws.cell(row=row, column=3).value = b or None
        ws.cell(row=row, column=4).value = c or None
        ws.cell(row=row, column=5).value = (c - b) or None
        ws.cell(row=row, column=6).value = safe_rate(c - b, b)

    # 依模板既有的「類別4代碼」(A 欄) 逐列對應，動態讀取避免硬編列號
    sa_row = None
    pure_b = pure_c = 0
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if code is None:
            continue
        code_f = float(code)
        if code_f == 4040.0:        # SA CARE
            b, c = sa_rev_total(df_prev, prv_s, prv_e), sa_rev_total(df_cur, cur_s, cur_e)
            sa_row = r
        else:                        # 一般 3PP 類別
            b, c = c4_rev(df_prev, prv_s, prv_e, code_f), c4_rev(df_cur, cur_s, cur_e, code_f)
            pure_b += b
            pure_c += c
        write_row(r, b, c)

    sa_b = sa_rev_total(df_prev, prv_s, prv_e)
    sa_c = sa_rev_total(df_cur, cur_s, cur_e)
    # 加總（純配件 + SA CARE）與 純配件：依 B 欄文字找列
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if label == '加總':
            write_row(r, pure_b + sa_b, pure_c + sa_c)
        elif label == '純配件':
            write_row(r, pure_b, pure_c)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    # --version 模式：只印版本號後離開
    if args.version:
        print(VERSION)
        sys.exit(0)

    # 確認必要參數都有傳入
    for attr in ['week_start', 'week_end', 'data_dir', 'template']:
        if not getattr(args, attr):
            sys.exit(f'ERROR: 缺少必要參數 --{attr.replace("_","-")}')
    if not args.output and not args.output_dir:
        sys.exit('ERROR: 缺少必要參數 --output 或 --output-dir')

    wk_start = date.fromisoformat(args.week_start)
    wk_end   = date.fromisoformat(args.week_end)
    data_dir = Path(args.data_dir).resolve()
    template = Path(args.template).resolve()

    # 解析輸出路徑（若提供 --output-dir，自動產生帶季週次的檔名）
    if args.output:
        output = Path(args.output).resolve()
    else:
        out_dir = Path(args.output_dir).resolve()
        out_dir.mkdir(parents=True, exist_ok=True)
        _wb_tmp = openpyxl.load_workbook(template, read_only=True, data_only=True)
        _fys = load_fiscal_year_start(_wb_tmp)
        _wb_tmp.close()
        if _fys:
            _q, _w = compute_fiscal_week_number(wk_start, _fys)
            fname = f'Q{_q}W{_w:02d}_週報_{wk_start}~{wk_end.strftime("%m-%d")}.xlsx'
        else:
            fname = f'週報_{wk_start}~{wk_end.strftime("%m-%d")}.xlsx'
        output = out_dir / fname

    print(f'=== 週報 Excel 填充  v{VERSION} ===')
    print(f'  本週: {wk_start} ~ {wk_end}')

    # ── 期間計算 ─────────────────────────────────────────────────────────────
    year, prev_year = wk_start.year, wk_start.year - 1
    cross_month = wk_end.month != wk_start.month

    report_month = wk_start.month
    prev_month   = report_month - 1 if report_month > 1 else 12
    prev_mo_year = year if report_month > 1 else year - 1

    if cross_month:
        mo_start = date(year, report_month, 1)
        mo_end   = date(year, report_month, monthrange(year, report_month)[1])
        lm_start = date(prev_mo_year, prev_month, 1)
        lm_end   = date(prev_mo_year, prev_month, monthrange(prev_mo_year, prev_month)[1])
        ly_start = date(prev_year, report_month, 1)
        ly_end   = date(prev_year, report_month, monthrange(prev_year, report_month)[1])
    else:
        mo_start = date(year, report_month, 1)
        mo_end   = wk_end
        lm_last  = monthrange(prev_mo_year, prev_month)[1]
        lm_start = date(prev_mo_year, prev_month, 1)
        lm_end   = date(prev_mo_year, prev_month, min(wk_end.day, lm_last))
        ly_last  = monthrange(prev_year, report_month)[1]
        ly_start = date(prev_year, report_month, 1)
        ly_end   = date(prev_year, report_month, min(wk_end.day, ly_last))

    prev_wk_end   = wk_start - timedelta(days=1)
    prev_wk_start = prev_wk_end - timedelta(days=6)
    # quarter_start is computed after loading template (needs fiscal_year_start from 設定 sheet)

    dates = dict(
        wk_start=wk_start, wk_end=wk_end,
        prev_wk_start=prev_wk_start, prev_wk_end=prev_wk_end,
        mo_start=mo_start, mo_end=mo_end,
        lm_start=lm_start, lm_end=lm_end,
        ly_start=ly_start, ly_end=ly_end,
        ytd_cur_start=date(year, 1, 1),
        ytd_prv_start=date(prev_year, 1, 1),
    )
    if args.yoy_end:
        dates['yoy_end'] = date.fromisoformat(args.yoy_end)
        print(f'  年對年截止日(自訂): {dates["yoy_end"]}')

    print(f'  本月: {mo_start} ~ {mo_end}')
    print(f'  上月: {lm_start} ~ {lm_end}')
    print(f'  去年: {ly_start} ~ {ly_end}')

    # ── 找資料檔 ─────────────────────────────────────────────────────────────
    file_cur  = find_800ab(data_dir, year)
    file_prev = find_800ab(data_dir, prev_year)
    if not file_cur:  sys.exit(f'ERROR: 找不到 800AB_{year} 檔案')
    if not file_prev: sys.exit(f'ERROR: 找不到 800AB_{prev_year}整年 檔案')

    sacare_path = find_sacare(data_dir)
    if not sacare_path: sys.exit('ERROR: 找不到 SAcare對應價目表.xlsx')

    print(f'\n  本年: {Path(file_cur).name}')
    print(f'  去年: {Path(file_prev).name}')

    # ── 載入資料 ─────────────────────────────────────────────────────────────
    print('\nLoading data...', flush=True)
    sacare_prices = load_sacare(sacare_path)
    df_cur  = load_800(file_cur)
    df_prev = load_800(file_prev)

    # ── 開啟模板 ─────────────────────────────────────────────────────────────
    print('\nFilling sheets...', flush=True)
    shutil.copy(template, output)
    wb = openpyxl.load_workbook(output)

    # 從模板的「設定」工作表動態讀取員工清單 & 年度起始日
    global EMPLOYEES, EMP_CODES
    loaded = load_employees(wb)
    if loaded:
        EMPLOYEES = loaded
        EMP_CODES  = [e[0] for e in EMPLOYEES]
        print(f'  員工設定: 從模板「設定」工作表讀取，共 {len(EMPLOYEES)} 人')
    else:
        print(f'  員工設定: 未找到「設定」工作表，使用預設清單（{len(EMPLOYEES)} 人）')

    fiscal_year_start = load_fiscal_year_start(wb)
    if fiscal_year_start:
        q_num, w_num = compute_fiscal_week_number(wk_start, fiscal_year_start)
        print(f'  財務週次: 年度起始 {fiscal_year_start}，本週為 Q{q_num} W{w_num:02d}')
    else:
        print(f'  財務週次: 未設定年度起始日，使用曆法季度')

    quarter_start = compute_quarter_start(wk_start, fiscal_year_start)

    fill_sheet1(wb['1.主機銷售台數'], df_cur, quarter_start, wk_end)
    fill_sheet2(wb['2.門市週報 '],   df_cur, df_prev, sacare_prices, dates)
    fill_sheet_focus(wb['3.每月重點'], df_cur, dates)
    fill_sheet_speakers(wb['4.Speakers'], df_cur, dates)
    fill_sheet3(wb['5.3PP配件比較'], df_cur, df_prev, sacare_prices, dates)
    fill_sheet45(wb['6.3PP 銷售排名'], wb['7.VAP銷售排名'], df_cur, sacare_prices, dates)
    fill_sheet6(wb['10.個人新制獎金'], df_cur, sacare_prices, dates)
    fill_sheet78(wb['11.個人週主機'], wb['12.個人月主機'], df_cur, sacare_prices, dates)
    fill_sheet9(wb['13.個人月3PP'], df_cur, sacare_prices, dates)
    fill_sheet10(wb['14.月報YOY'], df_cur, df_prev, sacare_prices, dates)
    fill_sheet11(wb['15.3PP YOY'], df_cur, df_prev, sacare_prices, dates)

    wb.save(output)
    print(f'\n✓ 完成: {output}')

if __name__ == '__main__':
    main()
